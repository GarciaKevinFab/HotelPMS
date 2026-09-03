"""
ZenStay - Sistema multi-hotel de gestion hotelera (PMS)
Idioma: espanol (Peru), moneda: PEN, zona horaria: America/Lima

Base de datos: Postgres (ver db/schema.sql y backend/db_pg.py). Antes era
MongoDB Atlas; se migro porque el cluster dejo de existir y porque los importes
en float no sirven para un sistema que emite comprobantes y cuadra caja.
"""

from fastapi import FastAPI, APIRouter, Depends, HTTPException, Header, Query, Body, Request
from fastapi.responses import StreamingResponse, FileResponse, RedirectResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import re
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any, Literal
import uuid
from datetime import datetime, timezone, date, timedelta
from enum import Enum
import bcrypt
import jwt
import base64
import json
import io
import asyncio
import plantillas_correo
import secrets
import hashlib

import db_pg
import checkout
import izipay

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# --------------------------------------------------------------------------
# Configuracion que NO puede tener valor por defecto
# --------------------------------------------------------------------------
# Antes JWT_SECRET caia a un placeholder ('...-change-in-production') si la
# variable faltaba. Eso es peor que no arrancar: el sistema levantaba y firmaba
# tokens con un secreto que esta escrito en el repositorio, asi que cualquiera
# podia fabricarse un token de SUPER_ADMIN. Ahora falta la variable y el
# proceso no arranca, que es ruidoso pero honesto.
JWT_SECRET = os.environ.get('JWT_SECRET')
if not JWT_SECRET:
    raise RuntimeError(
        "JWT_SECRET no esta configurada. Es la clave con la que se firman los "
        "tokens de sesion: sin ella el backend no puede arrancar. Generar una "
        "con `openssl rand -hex 32` y ponerla en el .env del VPS."
    )
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

# Mismo criterio para CORS. Antes caia a '*' y, combinado con
# allow_credentials=True, era a la vez invalido para los navegadores e inseguro.
CORS_ORIGINS = [o.strip() for o in os.environ.get('CORS_ORIGINS', '').split(',') if o.strip()]
if not CORS_ORIGINS:
    raise RuntimeError(
        "CORS_ORIGINS no esta configurada. Debe listar los origenes exactos "
        "permitidos, separados por coma (ej. https://zenstay.sisac.pe). No se "
        "acepta '*' porque las peticiones van con credenciales."
    )

# App Configuration
app = FastAPI(title="ZenStay API", version="1.0.0")
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============== ENUMS ==============
class Role(str, Enum):
    SUPER_ADMIN = "SUPER_ADMIN"
    ADMIN = "ADMIN"
    RECEPTIONIST = "RECEPTIONIST"
    HOUSEKEEPING = "HOUSEKEEPING"
    SECURITY = "SECURITY"

class OccupancyStatus(str, Enum):
    VACANT = "VACANT"
    OCCUPIED = "OCCUPIED"
    DUE_OUT = "DUE_OUT"

class HousekeepingStatus(str, Enum):
    DIRTY = "DIRTY"
    CLEANING = "CLEANING"
    CLEAN = "CLEAN"
    INSPECT = "INSPECT"
    OUT_OF_ORDER = "OUT_OF_ORDER"

class ReservationStatus(str, Enum):
    PREBOOK = "PREBOOK"
    CONFIRMED = "CONFIRMED"
    CHECKED_IN = "CHECKED_IN"
    CHECKED_OUT = "CHECKED_OUT"
    NO_SHOW = "NO_SHOW"
    CANCELLED = "CANCELLED"

class InvoiceType(str, Enum):
    BOLETA = "BOLETA"
    FACTURA = "FACTURA"

class InvoiceStatus(str, Enum):
    DRAFT = "DRAFT"
    SENT = "SENT"
    ACCEPTED = "ACCEPTED"
    REJECTED = "REJECTED"
    VOIDED = "VOIDED"
    PENDING = "PENDING"

class AlertSeverity(str, Enum):
    INFO = "INFO"
    WARN = "WARN"
    CRITICAL = "CRITICAL"

class PaymentMethod(str, Enum):
    CASH = "EFECTIVO"
    CARD = "TARJETA"
    TRANSFER = "TRANSFERENCIA"
    YAPE_PLIN = "YAPE_PLIN"
    OTHER = "OTRO"

class DocType(str, Enum):
    DNI = "DNI"
    CE = "CE"
    PASSPORT = "PASAPORTE"
    RUC = "RUC"

# ============== PYDANTIC MODELS ==============
# PyObjectId y serialize_doc existian para lidiar con los ObjectId de Mongo.
# Con Postgres las claves ya son uuid y la columna ya se llama `id`, asi que el
# renombrado de `_id` desaparece. La conversion de tipos para la respuesta JSON
# (uuid -> str, timestamptz -> ISO, numeric -> float) vive ahora en
# db_pg.to_api(). Se deja este alias porque hay endpoints que arman diccionarios
# a mano y lo llaman.
serialize_doc = db_pg.to_api


def id_valido(valor, campo="id"):
    """Valida un identificador que llega del cliente y devuelve 400 si no lo es.

    Ocupa el lugar de ObjectId(id): antes, un id con formato invalido reventaba
    dentro del driver. Aca se traduce a una respuesta clara.
    """
    try:
        return db_pg.a_uuid(valor, campo)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

# Auth Models
class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    full_name: str
    role: Role
    tenant_id: Optional[str] = None

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict

# Tenant Models
class TenantCreate(BaseModel):
    name: str
    ruc: str
    razon_social: Optional[str] = None
    nombre_comercial: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    # Admin user
    admin_email: Optional[str] = None
    admin_password: Optional[str] = None
    admin_name: Optional[str] = None

class TenantInvoicingConfig(BaseModel):
    """Configuracion de facturacion. TODOS los campos son opcionales.

    Antes tenian valores por defecto (B001, correlativo 1, IGV 18). El frontend
    manda solo la ruta y el token de NubeFact, asi que Pydantic rellenaba el
    resto con esos defaults y el UPDATE devolvia el correlativo de boletas a 1:
    guardar el token de facturacion reiniciaba la numeracion, y SUNAT no admite
    que una serie repita numeros.

    Ahora lo que no viene se queda como esta (ver el coalesce del endpoint).
    Los valores iniciales los pone la tabla en db/schema.sql, no este modelo.
    """
    nubefact_ruta: Optional[str] = None
    nubefact_token: Optional[str] = None
    invoicing_mode: Optional[Literal["MOCK", "LIVE"]] = None
    boleta_series: Optional[str] = None
    boleta_correlative: Optional[int] = None
    factura_series: Optional[str] = None
    factura_correlative: Optional[int] = None
    igv_rate: Optional[float] = None

# Room Models
class RoomTypeCreate(BaseModel):
    name: str
    capacity: int
    amenities: List[str] = []
    base_price: float

class RoomCreate(BaseModel):
    number: str
    floor: int
    room_type_id: str
    notes: Optional[str] = None

class RoomBulkCreate(BaseModel):
    room_type_id: str
    floor: int
    start_number: int
    count: int
    prefix: str = ""

# Guest Models
class GuestCreate(BaseModel):
    doc_type: DocType
    doc_number: str
    full_name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    nationality: Optional[str] = "PE"
    address: Optional[str] = None

# Reservation Models
class ReservationCreate(BaseModel):
    guest_id: str
    checkin_date: date
    checkout_date: date
    room_type_id: str
    room_id: Optional[str] = None
    adults: int = 1
    children: int = 0
    total_estimated: float
    deposit_amount: float = 0
    source: str = "DIRECTO"
    notes: Optional[str] = None

class ReservationUpdate(BaseModel):
    checkin_date: Optional[date] = None
    checkout_date: Optional[date] = None
    room_id: Optional[str] = None
    status: Optional[ReservationStatus] = None
    notes: Optional[str] = None
    cancel_reason: Optional[str] = None

# Group Reservation Models
class GroupReservationCreate(BaseModel):
    group_name: str
    contact_name: str
    contact_phone: Optional[str] = None
    contact_email: Optional[str] = None
    checkin_date: date
    checkout_date: date
    rooms: List[dict]  # [{"room_type_id": "...", "quantity": 2}]
    adults: int = 1
    children: int = 0
    deposit_amount: float = 0
    notes: Optional[str] = None

# Email Models
class EmailRequest(BaseModel):
    recipient_email: EmailStr
    subject: str
    html_content: str

class EmailTemplate(str, Enum):
    RESERVATION_CONFIRMATION = "RESERVATION_CONFIRMATION"
    CHECKIN_CONFIRMATION = "CHECKIN_CONFIRMATION"
    CHECKOUT_REMINDER = "CHECKOUT_REMINDER"
    PAYMENT_RECEIPT = "PAYMENT_RECEIPT"

# Rate Management Models
class RateCreate(BaseModel):
    room_type_id: str
    date_from: date
    date_to: date
    price: float
    name: Optional[str] = None  # e.g., "Temporada Alta", "Feriado"
    min_stay: int = 1

# Folio Models
class ChargeCreate(BaseModel):
    concept: str
    quantity: float = 1
    unit_price: float
    tax_type: str = "IGV"
    category: str = "HABITACION"

class PaymentCreate(BaseModel):
    method: PaymentMethod
    amount: float
    reference: Optional[str] = None

class VoidRequest(BaseModel):
    reason: str

# Cash Shift Models
class CashShiftOpen(BaseModel):
    opening_amount: float

class CashShiftClose(BaseModel):
    counted_cash: float
    notes: Optional[str] = None

class CashMovementCreate(BaseModel):
    type: Literal["IN", "OUT"]
    amount: float
    reason: str

# Invoice Models
class InvoiceCreate(BaseModel):
    folio_id: str
    type: InvoiceType
    client_doc_type: DocType
    client_doc_number: str
    client_name: str
    client_address: Optional[str] = None

# Housekeeping Models
class HousekeepingStatusUpdate(BaseModel):
    status: HousekeepingStatus
    notes: Optional[str] = None

# Maintenance Models
class MaintenanceTicketCreate(BaseModel):
    room_id: str
    title: str
    description: str
    priority: Literal["LOW", "MEDIUM", "HIGH", "CRITICAL"]
    estimated_cost: Optional[float] = None

# Alert Models
class AlertResolve(BaseModel):
    notes: Optional[str] = None

# ============== AUTH HELPERS ==============
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token(user_id: str, email: str, role: str, tenant_id: str = None, extra: dict = None) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "tenant_id": tenant_id,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    # `extra` lleva las marcas de "SUPER_ADMIN dentro de un hotel" (ver
    # POST /tenants/{id}/entrar): en_otro_hotel y hotel_nombre. Van en el
    # token y no en la base porque describen ESTA sesion, no al usuario.
    if extra:
        payload.update(extra)
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def es_superadmin(user: dict) -> bool:
    """SUPER_ADMIN de verdad, este o no dentro de un hotel.

    Cuando entra a un hotel (POST /tenants/{id}/entrar) su token lleva rol
    ADMIN para que RLS y los permisos lo traten como uno mas del hotel; pero
    las cosas que un SUPER_ADMIN nunca sufre -la suscripcion suspendida, el
    cupo de habitaciones- tampoco tienen que frenarlo ahi dentro: entro
    justamente a arreglarlas.
    """
    return user.get("role") == Role.SUPER_ADMIN.value or bool(user.get("en_otro_hotel"))

async def get_current_user(authorization: str = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Token no proporcionado")
    token = authorization.replace("Bearer ", "")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

def require_roles(*roles: Role):
    async def role_checker(user: dict = Depends(get_current_user)):
        if user["role"] not in [r.value for r in roles]:
            raise HTTPException(status_code=403, detail="Permisos insuficientes")
        return user
    return role_checker

# ============== SUSCRIPCIONES ==============
# Ver db/migrations/002_suscripciones.sql para los estados y por que existe la
# gracia. Aqui vive la parte que se consulta en caliente.

DIAS_PRUEBA = 14
DIAS_GRACIA = 7

# `vencida` SIGUE teniendo acceso: es la gracia. Solo 'suspendida' y
# 'cancelada' cortan. Un hotel con huespedes dentro no puede quedarse sin poder
# cobrarles porque una tarjeta rebotara anoche.
ESTADOS_CON_ACCESO = ("prueba", "activa", "vencida")


async def estado_suscripcion(conn, tenant_id):
    """Suscripcion del hotel + su plan + si tiene acceso, resuelto en un sitio."""
    fila = await db_pg.uno(
        conn,
        """select t.subscription_status, t.trial_ends_at, t.grace_ends_at,
                  p.codigo as plan_codigo, p.nombre as plan_nombre,
                  p.precio_mensual, p.max_habitaciones,
                  p.facturacion_sunat, p.reportes_avanzados,
                  (select count(*) from rooms where tenant_id = t.id and is_active) as habitaciones_usadas
           from tenants t
           join planes p on p.codigo = t.plan_codigo
           where t.id = $1""",
        db_pg.a_uuid(tenant_id),
    )
    if not fila:
        return None
    fila["tiene_acceso"] = fila["subscription_status"] in ESTADOS_CON_ACCESO
    return fila


async def exigir_suscripcion(user: dict = Depends(get_current_user)):
    """Corta el paso a los hoteles suspendidos o cancelados.

    Va como dependencia en los endpoints que ESCRIBEN, no en los que leen: a un
    hotel suspendido se le deja consultar y exportar sus datos -- son suyos --
    pero no seguir operando. Cortar tambien la lectura convierte un impago en
    una perdida de informacion, que no es lo que se acordo con el cliente.

    El SUPER_ADMIN nunca se bloquea: es quien tiene que poder entrar a
    arreglarlo.
    """
    if es_superadmin(user):
        return user

    async with db_pg.tx(user) as conn:
        susc = await estado_suscripcion(conn, user["tenant_id"])

    if susc and not susc["tiene_acceso"]:
        raise HTTPException(
            status_code=402,  # Payment Required: dice exactamente lo que pasa
            detail="La suscripción de este hotel está suspendida. "
                   "Renuévala para seguir operando."
        )
    return user


async def exigir_cupo_habitaciones(conn, user: dict, cuantas: int = 1):
    """Comprueba que el plan del hotel admite `cuantas` habitaciones mas.

    Se llama al CREAR, no al pintar la pantalla: ocultar el boton de "nueva
    habitacion" no impide que alguien mande el formulario igual, y el limite es
    justamente lo que se esta vendiendo.

    Cuenta solo las activas: una habitacion dada de baja no ocupa cupo, o dar de
    baja y volver a crear seria una forma tonta de saltarse el plan... al reves,
    seria injusto cobrar por habitaciones que el hotel ya no usa.
    """
    if es_superadmin(user):
        return

    limite = await db_pg.uno(
        conn,
        """select p.max_habitaciones, p.nombre,
                  (select count(*) from rooms where tenant_id = t.id and is_active) as usadas
           from tenants t join planes p on p.codigo = t.plan_codigo
           where t.id = $1""",
        db_pg.a_uuid(user["tenant_id"]),
    )
    # max_habitaciones NULL = sin limite (planes Prueba y Empresa).
    if not limite or limite["max_habitaciones"] is None:
        return

    if limite["usadas"] + cuantas > limite["max_habitaciones"]:
        raise HTTPException(
            status_code=402,
            detail=(
                f"El plan {limite['nombre']} permite hasta "
                f"{limite['max_habitaciones']} habitaciones y ya tienes "
                f"{limite['usadas']}. Cambia de plan para agregar más."
            ),
        )


def tenant_de(user: dict):
    """El hotel del usuario, o None si es SUPER_ADMIN (ve todos).

    Reemplaza a get_tenant_filter(), que devolvia un diccionario de filtro de
    Mongo. Ahora devuelve un valor suelto que se pasa como parametro a la
    consulta, siempre con la misma forma:

        where ($1::uuid is null or tenant_id = $1)

    Con un usuario normal, $1 es su hotel y filtra. Con un SUPER_ADMIN, $1 es
    NULL y la condicion se cumple para todas las filas -- exactamente lo que
    hacia el `return {}` de antes. Ademas RLS respalda esto por debajo: si a una
    consulta se le olvida la condicion, la politica devuelve cero filas en vez
    de datos de otro cliente.
    """
    if user["role"] == Role.SUPER_ADMIN.value:
        return None
    if not user.get("tenant_id"):
        raise HTTPException(status_code=400, detail="Usuario sin tenant asignado")
    return db_pg.a_uuid(user["tenant_id"], "tenant_id")


# ============== AUDIT HELPER ==============
async def create_audit_log(conn, tenant_id, user_id, entity: str, action: str, before: dict = None, after: dict = None):
    """Registra un cambio en la bitacora de auditoria.

    Recibe la conexion en vez de abrir una propia: asi la auditoria entra en la
    MISMA transaccion que el cambio que describe. Antes eran dos escrituras
    sueltas, y si la segunda fallaba quedaba un cambio sin rastro (o al reves,
    un rastro de algo que no llego a pasar).
    """
    await conn.execute(
        """insert into audit_logs (tenant_id, user_id, entity, action, before_json, after_json)
           values ($1, $2, $3, $4, $5, $6)""",
        db_pg.a_uuid(tenant_id), db_pg.a_uuid(user_id), entity, action,
        json.dumps(before) if before is not None else None,
        json.dumps(after) if after is not None else None,
    )

# ============== ALERT HELPER ==============
async def create_alert(conn, tenant_id, alert_type: str, severity: AlertSeverity, title: str, message: str, entity_ref: dict = None):
    """Crea una alerta operativa. Tambien participa de la transaccion que la origina."""
    await conn.execute(
        """insert into alerts (tenant_id, type, severity, title, message, entity_ref, status)
           values ($1, $2, $3::alert_severity, $4, $5, $6, 'OPEN')""",
        db_pg.a_uuid(tenant_id), alert_type, severity.value, title, message,
        json.dumps(entity_ref or {}),
    )

# ============== NUBEFACT MOCK SERVICE ==============
class NubeFactService:
    @staticmethod
    async def send_invoice(tenant: dict, invoice_data: dict) -> dict:
        config = tenant.get("invoicing_config", {})
        mode = config.get("invoicing_mode", "MOCK")
        
        if mode == "MOCK":
            return NubeFactService._mock_response(invoice_data)
        else:
            return await NubeFactService._live_request(config, invoice_data)
    
    @staticmethod
    def _mock_response(invoice_data: dict) -> dict:
        series = invoice_data.get("serie", "B001")
        number = invoice_data.get("numero", 1)
        
        mock_pdf = f"COMPROBANTE ELECTRONICO\nSerie: {series}\nNumero: {number}\nCliente: {invoice_data.get('cliente_denominacion')}\nTotal: S/ {invoice_data.get('total', 0):.2f}".encode()
        mock_xml = f'<?xml version="1.0"?><Invoice><Series>{series}</Series><Number>{number}</Number></Invoice>'.encode()
        mock_cdr = f'<?xml version="1.0"?><CDR><Status>Aceptado</Status></CDR>'.encode()
        
        return {
            "success": True,
            "aceptada_por_sunat": True,
            "serie": series,
            "numero": number,
            "enlace_del_pdf": f"https://mock.nubefact.com/pdf/{series}-{number}.pdf",
            "enlace_del_xml": f"https://mock.nubefact.com/xml/{series}-{number}.xml",
            "enlace_del_cdr": f"https://mock.nubefact.com/cdr/{series}-{number}.xml",
            "pdf_base64": base64.b64encode(mock_pdf).decode(),
            "xml_base64": base64.b64encode(mock_xml).decode(),
            "cdr_base64": base64.b64encode(mock_cdr).decode(),
            "hash": f"MOCK-HASH-{series}-{number}",
            "qr": f"https://mock.nubefact.com/qr/{series}-{number}",
            "sunat_description": "La operación se realizó satisfactoriamente (MOCK)",
            "sunat_responsecode": "0"
        }
    
    @staticmethod
    async def _live_request(config: dict, invoice_data: dict) -> dict:
        import requests
        ruta = config.get("nubefact_ruta")
        token = config.get("nubefact_token")
        
        if not ruta or not token:
            return {"success": False, "error": "Configuración NubeFact incompleta"}
        
        headers = {
            "Authorization": f'Token token="{token}"',
            "Content-Type": "application/json"
        }
        
        try:
            response = requests.post(ruta, json=invoice_data, headers=headers, timeout=30)
            result = response.json()
            result["success"] = result.get("aceptada_por_sunat", False)
            return result
        except Exception as e:
            return {"success": False, "error": str(e)}

# ============== AUTH ENDPOINTS ==============
@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    # db_pg.autenticar() es la unica lectura que ocurre sin saber el hotel:
    # averiguarlo es justamente lo que esta haciendo. Por debajo llama a la
    # funcion app_autenticar() de Postgres, acotada a una fila y a las columnas
    # del login (ver db/rls.sql).
    user = await db_pg.autenticar(credentials.email)
    if not user:
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    if not user.get("is_active", True):
        raise HTTPException(status_code=401, detail="Usuario desactivado")
    if not verify_password(credentials.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")

    # Un hotel desactivado por el SUPER_ADMIN no entra, con ninguna cuenta.
    # Se comprueba DESPUES de la contrasena para no revelar el estado del
    # hotel a quien no la sabe. El SUPER_ADMIN no tiene hotel y nunca se
    # bloquea: es quien tiene que poder entrar a reactivarlo.
    if user.get("tenant_id") and user["role"] != Role.SUPER_ADMIN.value:
        async with db_pg.tx(user) as conn:
            hotel_activo = await db_pg.valor(
                conn, "select is_active from tenants where id = $1",
                db_pg.a_uuid(user["tenant_id"], "tenant_id"),
            )
        if hotel_activo is False:
            raise HTTPException(
                status_code=403,
                detail="Este hotel está desactivado. Escribe a soporte@sisac.pe",
            )

    token = create_token(user["id"], user["email"], user["role"], user.get("tenant_id"))

    return TokenResponse(
        access_token=token,
        user={
            "id": user["id"],
            "email": user["email"],
            "full_name": user["full_name"],
            "role": user["role"],
            "tenant_id": user.get("tenant_id")
        }
    )

@api_router.get("/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    en_otro_hotel = bool(user.get("en_otro_hotel"))
    # Dentro de un hotel el token lleva el tenant del hotel, pero la fila del
    # SUPER_ADMIN tiene tenant_id NULL: bajo el aislamiento del hotel no se ve.
    # Se lee con la puerta global, acotada a UN id que ya viene firmado en el
    # token.
    contexto = (
        db_pg.tx_global("superadmin dentro de un hotel: su propia fila no tiene tenant")
        if en_otro_hotel else db_pg.tx(user)
    )
    async with contexto as conn:
        db_user = await db_pg.uno(
            conn,
            """select id, tenant_id, email, full_name, role, is_active, created_at
               from users where id = $1""",
            id_valido(user["user_id"], "user_id"),
        )
        if not db_user:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")

        # El nombre del hotel, para la cabecera de la UI. Un SUPER_ADMIN no
        # tiene hotel propio, asi que no se consulta.
        if user.get("tenant_id"):
            db_user["tenant"] = await db_pg.uno(
                conn,
                "select name, nombre_comercial from tenants where id = $1",
                id_valido(user["tenant_id"], "tenant_id"),
            )

    if en_otro_hotel:
        # La UI se comporta como la de un ADMIN del hotel, y ademas sabe que
        # en realidad es el SUPER_ADMIN de visita (para la franja de aviso).
        db_user["rol_real"] = db_user["role"]
        db_user["role"] = Role.ADMIN.value
        db_user["tenant_id"] = user.get("tenant_id")
        db_user["en_otro_hotel"] = True
        db_user["hotel_nombre"] = user.get("hotel_nombre")

    return db_user


class CambioPassword(BaseModel):
    actual: str
    nueva: str = Field(min_length=8, max_length=128)


@api_router.put("/auth/password")
async def cambiar_mi_password(data: CambioPassword, user: dict = Depends(get_current_user)):
    """Cualquier usuario cambia SU contrasena sabiendo la actual.

    Hasta ahora solo existia el restablecimiento por un ADMIN, y el
    SUPER_ADMIN -que no tiene ningun ADMIN por encima- no tenia forma de
    cambiar la suya sin tocar la base.
    """
    if user.get("en_otro_hotel"):
        raise HTTPException(status_code=400, detail="Sal del hotel antes de cambiar tu contraseña")

    uid = id_valido(user["user_id"], "user_id")
    # Puerta global a proposito: el usuario se identifica por el id firmado en
    # su token, y el SUPER_ADMIN no tiene tenant con el que abrir tx().
    async with db_pg.tx_global("cambiar la propia contrasena: fila unica por id del token") as conn:
        fila = await db_pg.uno(conn, "select id, tenant_id, password_hash from users where id = $1", uid)
        if not fila:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        if not verify_password(data.actual, fila["password_hash"]):
            raise HTTPException(status_code=400, detail="La contraseña actual no es correcta")
        await conn.execute("update users set password_hash = $2 where id = $1", uid, hash_password(data.nueva))
        if fila.get("tenant_id"):
            await create_audit_log(conn, fila["tenant_id"], uid, "user", "PASSWORD_CHANGED_SELF",
                                   None, {"user_id": str(uid)})
    return {"message": "Contraseña actualizada"}


class PerfilPropio(BaseModel):
    full_name: str = Field(min_length=2, max_length=120)


@api_router.put("/auth/perfil")
async def actualizar_mi_perfil(data: PerfilPropio, user: dict = Depends(get_current_user)):
    """El nombre con el que uno aparece en la app. Solo eso: el correo es la
    identidad de acceso y lo cambia un ADMIN; el rol, ni hablar."""
    if user.get("en_otro_hotel"):
        raise HTTPException(status_code=400, detail="Sal del hotel antes de editar tu cuenta")
    uid = id_valido(user["user_id"], "user_id")
    async with db_pg.tx_global("editar el propio nombre: fila unica por id del token") as conn:
        actualizado = await conn.fetchval(
            "update users set full_name = $2 where id = $1 returning id", uid, data.full_name.strip()
        )
        if not actualizado:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return {"message": "Nombre actualizado", "full_name": data.full_name.strip()}


@api_router.post("/auth/salir-de-hotel")
async def salir_de_hotel(user: dict = Depends(get_current_user)):
    """Devuelve al SUPER_ADMIN su token global despues de visitar un hotel.

    Es un endpoint propio, y no "entra a tu propio hotel", porque el camino
    de vuelta tiene que funcionar aunque el hotel visitado se haya borrado o
    desactivado mientras tanto.
    """
    if not user.get("en_otro_hotel"):
        raise HTTPException(status_code=400, detail="No estás dentro de ningún hotel")

    uid = id_valido(user["user_id"], "user_id")
    async with db_pg.tx_global("salir de un hotel: comprobar que el id del token sigue siendo SUPER_ADMIN") as conn:
        fila = await db_pg.uno(
            conn, "select id, email, role, is_active from users where id = $1", uid
        )
    if not fila or fila["role"] != Role.SUPER_ADMIN.value or not fila.get("is_active", True):
        raise HTTPException(status_code=403, detail="Esta sesión ya no pertenece a un Super Admin")

    token = create_token(fila["id"], fila["email"], Role.SUPER_ADMIN.value, None)
    return {"access_token": token, "token_type": "bearer", "message": "De vuelta en la consola"}

# ============== PLANES Y ALTA PUBLICA ==============
class RegistroHotel(BaseModel):
    """Alta desde la landing. Sin SUPER_ADMIN de por medio."""
    hotel_name: str = Field(min_length=2, max_length=120)
    ruc: str = Field(min_length=11, max_length=11)
    admin_name: str = Field(min_length=2, max_length=120)
    admin_email: EmailStr
    admin_password: str = Field(min_length=8)


@api_router.get("/planes")
async def listar_planes():
    """Catalogo publico. Lo pinta la landing sin que nadie haya entrado."""
    async with db_pg.tx_global("catalogo publico de planes, no depende de ningun hotel") as conn:
        return await db_pg.varias(
            conn,
            """select codigo, nombre, descripcion, precio_mensual,
                      max_habitaciones, facturacion_sunat, reportes_avanzados
               from planes where activo order by orden""",
        )


@api_router.post("/registro")
async def registro_publico(data: RegistroHotel):
    """Da de alta un hotel y su administrador, y arranca la prueba gratuita.

    Es el unico endpoint que crea hoteles sin ser SUPER_ADMIN, y por eso lleva
    las validaciones que POST /tenants se puede ahorrar:

    - El RUC tiene que ser once digitos. No se valida el digito verificador
      aqui a proposito: un RUC mal tecleado se corrige, pero rechazar un alta
      por una regla mal implementada pierde un cliente en la puerta.
    - La contrasena, minimo 8 caracteres (lo exige el modelo).
    - Hotel y administrador se crean en UNA transaccion: si el email ya existe,
      no queda un hotel huerfano sin nadie que pueda entrar.
    """
    if not data.ruc.isdigit():
        raise HTTPException(status_code=400, detail="El RUC debe ser numérico, de 11 dígitos")

    async with db_pg.tx_global("alta publica: el hotel todavia no existe") as conn:
        if await conn.fetchval("select 1 from tenants where ruc = $1", data.ruc):
            raise HTTPException(status_code=400, detail="Ya existe un hotel registrado con este RUC")
        if await conn.fetchval("select 1 from users where email = $1", data.admin_email):
            raise HTTPException(status_code=400, detail="Ese correo ya está registrado")

        tenant_id = await conn.fetchval(
            """insert into tenants (name, ruc, razon_social, nombre_comercial, email,
                                    plan_codigo, subscription_status, trial_ends_at)
               values ($1, $2, $1, $1, $3, 'prueba', 'prueba', now() + ($4 || ' days')::interval)
               returning id""",
            data.hotel_name, data.ruc, data.admin_email, str(DIAS_PRUEBA),
        )
        await conn.execute(
            """insert into users (tenant_id, email, password_hash, full_name, role)
               values ($1, $2, $3, $4, 'ADMIN')""",
            tenant_id, data.admin_email, hash_password(data.admin_password), data.admin_name,
        )

    return {
        "message": "Hotel registrado. Tu prueba gratuita ya está activa.",
        "tenant_id": str(tenant_id),
        "dias_prueba": DIAS_PRUEBA,
    }


@api_router.get("/suscripcion")
async def ver_suscripcion(user: dict = Depends(get_current_user)):
    """Estado de la suscripcion del hotel, para la barra de aviso y Ajustes."""
    if user.get("role") == Role.SUPER_ADMIN.value:
        return {"subscription_status": "activa", "plan_nombre": "Super Admin", "tiene_acceso": True}
    async with db_pg.tx(user) as conn:
        susc = await estado_suscripcion(conn, user["tenant_id"])
    if not susc:
        raise HTTPException(status_code=404, detail="Hotel no encontrado")
    return susc


# ============== CONFIRMACION DEL PAGO (IPN DE IZIPAY) ==============
# Cuantos dias suma cada periodo. Van aqui y no en la consulta para que el
# numero se lea una sola vez y no haya un 30 escondido dentro de un SQL.
DIAS_POR_PERIODO = {"mensual": 30, "anual": 365}


@api_router.post("/checkout/izipay/ipn")
async def izipay_ipn(request: Request):
    """Confirmacion servidor-a-servidor de Izipay. Publica y sin sesion.

    LA AUTENTICIDAD LA DA LA FIRMA, NO LA COOKIE

      Este aviso llega desde los servidores de la pasarela, sin navegador y sin
      sesion. Lo unico que distingue uno legitimo de uno que mande cualquiera
      desde internet es el HMAC, asi que sin firma valida se responde 401 y no
      se toca nada. izipay.verificar_firma() devuelve False cuando falta la
      clave: falla cerrado a proposito.

    ES IDEMPOTENTE, Y ESO NO ES UN DETALLE

      Las pasarelas reenvian. El `and estado <> 'pagado'` del UPDATE es lo que
      hace que la segunda notificacion del mismo pedido no encuentre fila que
      actualizar y, por tanto, no vuelva a extender el periodo. Sin eso el bug
      seria "a algunos hoteles se les regalan meses" y nadie lo reportaria.

    Va con tx_global porque no hay usuario del que heredar el hotel: el numero
    de orden es lo unico que dice a que hotel pertenece este pago.
    """
    crudo = await request.body()
    # POR_CONFIRMAR: el nombre de la cabecera que trae la firma. Se aceptan las
    # dos formas que aparecen en la documentacion.
    firma = (request.headers.get("x-izipay-signature")
             or request.headers.get("signature") or "")

    if not izipay.verificar_firma(crudo, firma):
        logger.warning("IPN de Izipay con firma invalida desde %s",
                       request.client.host if request.client else "?")
        raise HTTPException(status_code=401, detail="Firma inválida")

    try:
        datos = json.loads(crudo)
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="El cuerpo no es JSON")

    # POR_CONFIRMAR: los nombres exactos de los campos del aviso.
    respuesta = datos.get("response") if isinstance(datos.get("response"), dict) else {}
    numero = (datos.get("orderNumber") or datos.get("order_number")
              or respuesta.get("orderNumber"))
    if not numero:
        raise HTTPException(status_code=400, detail="Sin número de orden")

    aprobado = (str(datos.get("code")) == "00"
                or datos.get("status") in ("PAID", "AUTHORIZED"))
    if not aprobado:
        logger.info("IPN de %s indica pago no aprobado: %s", numero, datos.get("message"))
        # 200 igualmente: el aviso se recibio y se entendio. Un 4xx haria que
        # Izipay lo reintentara para siempre.
        return {"ok": True, "aplicado": False, "motivo": "pago no aprobado"}

    transaction_id = datos.get("transactionId") or respuesta.get("transactionId")

    async with db_pg.tx_global("IPN de Izipay: llega sin sesion y sin hotel conocido") as conn:
        pago = await db_pg.uno(
            conn,
            """update pagos_suscripcion
                  set estado = 'pagado', confirmado_en = now(),
                      izipay_transaction_id = $2,
                      respuesta = coalesce(respuesta, '{}'::jsonb) || $3::jsonb
                where izipay_order_number = $1 and estado <> 'pagado'
            returning id, tenant_id, plan_codigo, periodo, monto""",
            numero, transaction_id, json.dumps({"ipn": datos}),
        )
        if not pago:
            # O ya estaba confirmado, o el numero no existe. En los dos casos
            # se responde 200: no hay nada que reintentar.
            logger.info("IPN %s ya aplicado o pedido inexistente: no se repite", numero)
            return {"ok": True, "aplicado": False, "motivo": "ya aplicado o desconocido"}

        dias = DIAS_POR_PERIODO.get(pago["periodo"], 30)
        # greatest(...) para que renovar antes de tiempo SUME al periodo que
        # queda en vez de recortarlo, y coalesce por los hoteles que nunca
        # pagaron (subscription_ends_at NULL).
        # db_pg.uno() ya convirtio el uuid a texto al salir; para volver a
        # consultar por el hay que devolverlo a uuid o asyncpg no lo acepta.
        tid = db_pg.a_uuid(pago["tenant_id"], "tenant_id")
        antes = await db_pg.uno(
            conn,
            "select plan_codigo, subscription_status, subscription_ends_at "
            "from tenants where id = $1", tid)
        await conn.execute(
            """update tenants
                  set plan_codigo = $2,
                      subscription_status = 'activa',
                      subscription_ends_at =
                          greatest(coalesce(subscription_ends_at, now()), now())
                          + ($3 || ' days')::interval
                where id = $1""",
            tid, pago["plan_codigo"], str(dias),
        )
        despues = await db_pg.uno(
            conn,
            "select plan_codigo, subscription_status, subscription_ends_at "
            "from tenants where id = $1", tid)

        # user_id None: no lo hizo una persona de este sistema, lo hizo la
        # pasarela. Dejarlo en blanco es mas honesto que atribuirselo a alguien.
        await create_audit_log(
            conn, tid, None, "suscripcion_pago", "confirm",
            before=antes,
            after={**despues, "numero_orden": numero, "monto": pago["monto"],
                   "izipay_transaction_id": transaction_id},
        )

    logger.info("IPN: pago %s confirmado, hotel %s activo hasta %s",
                numero, pago["tenant_id"], despues.get("subscription_ends_at"))
    return {"ok": True, "aplicado": True}


# ============== TENANT ENDPOINTS ==============
@api_router.post("/tenants")
async def create_tenant(data: TenantCreate, user: dict = Depends(require_roles(Role.SUPER_ADMIN))):
    # Crear un hotel y su administrador es UNA operacion: si el email del admin
    # ya estaba en uso, antes quedaba el hotel creado y sin administrador, y
    # habia que limpiarlo a mano. Ahora las dos escrituras van en la misma
    # transaccion y o entran las dos o no entra ninguna.
    async with db_pg.tx_global("crear un hotel nuevo: aun no existe su tenant_id") as conn:
        if await conn.fetchval("select 1 from tenants where ruc = $1", data.ruc):
            raise HTTPException(status_code=400, detail="Ya existe un hotel con este RUC")

        if data.admin_email and await conn.fetchval(
            "select 1 from users where email = $1", data.admin_email
        ):
            raise HTTPException(status_code=400, detail="El email del administrador ya está en uso")

        # La config de facturacion ya no va anidada en un campo `invoicing_config`
        # ademas de plana en la raiz: en el esquema nuevo existe una sola vez,
        # con los valores por defecto que declara la tabla (B001/F001, IGV 18%).
        tenant_id = await conn.fetchval(
            """insert into tenants (name, ruc, razon_social, nombre_comercial,
                                    address, phone, email)
               values ($1, $2, $3, $4, $5, $6, $7)
               returning id""",
            data.name, data.ruc,
            data.razon_social or data.name,
            data.nombre_comercial or data.name,
            data.address, data.phone, data.email,
        )

        admin_id = None
        if data.admin_email and data.admin_password:
            admin_id = await conn.fetchval(
                """insert into users (tenant_id, email, password_hash, full_name, role)
                   values ($1, $2, $3, $4, 'ADMIN')
                   returning id""",
                tenant_id, data.admin_email,
                hash_password(data.admin_password),
                data.admin_name or "Administrador",
            )

        await create_audit_log(conn, tenant_id, user["user_id"], "tenant", "CREATE",
                               None, {"tenant_id": str(tenant_id), "name": data.name})

    return {
        "id": str(tenant_id),
        "admin_id": str(admin_id) if admin_id else None,
        "message": "Hotel creado exitosamente"
    }

@api_router.get("/tenants")
async def list_tenants(user: dict = Depends(require_roles(Role.SUPER_ADMIN))):
    async with db_pg.tx(user) as conn:
        # Dos conteos por hotel para la tabla de la consola. Son subconsultas
        # correlacionadas sobre indices por tenant_id: con decenas de hoteles
        # cuesta lo mismo que el select a secas.
        return await db_pg.varias(
            conn,
            """select t.*,
                      (select count(*) from rooms r where r.tenant_id = t.id and r.is_active) as habitaciones,
                      (select count(*) from users u where u.tenant_id = t.id) as usuarios
               from tenants t order by t.name""",
        )

@api_router.get("/tenants/{tenant_id}")
async def get_tenant(tenant_id: str, user: dict = Depends(get_current_user)):
    if user["role"] != Role.SUPER_ADMIN.value and user.get("tenant_id") != tenant_id:
        raise HTTPException(status_code=403, detail="Acceso denegado")

    async with db_pg.tx(user) as conn:
        tenant = await db_pg.uno(conn, "select * from tenants where id = $1",
                                 id_valido(tenant_id, "tenant_id"))
    if not tenant:
        raise HTTPException(status_code=404, detail="Hotel no encontrado")
    return tenant


class TenantUpdate(BaseModel):
    """Datos editables de un hotel. Todo opcional: lo que no viene no cambia.

    El plan, el estado de suscripcion y is_active NO estan aqui a proposito:
    tienen sus propios endpoints, solo para SUPER_ADMIN, con su propia
    auditoria. Asi un ADMIN puede editar la ficha de su hotel con este mismo
    modelo sin que exista la tentacion de colar un campo de mas.
    """
    name: Optional[str] = Field(default=None, min_length=2, max_length=120)
    razon_social: Optional[str] = Field(default=None, max_length=160)
    nombre_comercial: Optional[str] = Field(default=None, max_length=120)
    ruc: Optional[str] = Field(default=None, min_length=11, max_length=11)
    address: Optional[str] = Field(default=None, max_length=200)
    phone: Optional[str] = Field(default=None, max_length=40)
    email: Optional[EmailStr] = None
    checkin_time: Optional[str] = Field(default=None, pattern=r"^\d{2}:\d{2}$")
    checkout_time: Optional[str] = Field(default=None, pattern=r"^\d{2}:\d{2}$")


@api_router.put("/tenants/{tenant_id}")
async def update_tenant(
    tenant_id: str, data: TenantUpdate,
    user: dict = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
):
    if user["role"] != Role.SUPER_ADMIN.value and user.get("tenant_id") != tenant_id:
        raise HTTPException(status_code=403, detail="Acceso denegado")
    if data.ruc is not None and not data.ruc.isdigit():
        raise HTTPException(status_code=400, detail="El RUC debe ser numérico, de 11 dígitos")

    cambios = {k: v for k, v in data.model_dump().items() if v is not None}
    if not cambios:
        raise HTTPException(status_code=400, detail="No hay campos válidos para actualizar")

    tid = id_valido(tenant_id, "tenant_id")
    # tx(user): para el ADMIN la politica RLS de `tenants` solo deja pasar su
    # propia fila, asi que aunque el 403 de arriba fallara, el UPDATE no
    # tocaria otro hotel. Para el SUPER_ADMIN pasa todo.
    async with db_pg.tx(user) as conn:
        antes = await db_pg.uno(
            conn,
            """select name, razon_social, nombre_comercial, ruc, address, phone, email,
                      checkin_time, checkout_time from tenants where id = $1""",
            tid,
        )
        if not antes:
            raise HTTPException(status_code=404, detail="Hotel no encontrado")

        # Las columnas salen de las claves del modelo (lista blanca), los
        # valores viajan como parametros.
        asignaciones = [f"{col} = ${i}" for i, col in enumerate(cambios, start=2)]
        try:
            await conn.execute(
                f"update tenants set {', '.join(asignaciones)} where id = $1",
                tid, *cambios.values(),
            )
        except db_pg.UniqueViolationError:
            raise HTTPException(status_code=400, detail="Ya existe otro hotel con ese RUC")

        await create_audit_log(
            conn, tid, user["user_id"], "tenant", "UPDATE",
            before={k: antes.get(k) for k in cambios}, after=cambios,
        )
    return {"message": "Hotel actualizado"}


class TenantActivo(BaseModel):
    is_active: bool


@api_router.put("/tenants/{tenant_id}/activo")
async def set_tenant_activo(
    tenant_id: str, data: TenantActivo,
    user: dict = Depends(require_roles(Role.SUPER_ADMIN)),
):
    """Apaga o enciende un hotel entero. Desactivado, nadie de ese hotel puede
    iniciar sesion (ver POST /auth/login); sus datos se quedan intactos."""
    tid = id_valido(tenant_id, "tenant_id")
    async with db_pg.tx(user) as conn:
        antes = await db_pg.valor(conn, "select is_active from tenants where id = $1", tid)
        if antes is None:
            raise HTTPException(status_code=404, detail="Hotel no encontrado")
        await conn.execute("update tenants set is_active = $2 where id = $1", tid, data.is_active)
        await create_audit_log(conn, tid, user["user_id"], "tenant",
                               "ACTIVATE" if data.is_active else "DEACTIVATE",
                               {"is_active": antes}, {"is_active": data.is_active})
    return {"message": "Hotel activado" if data.is_active else "Hotel desactivado", "is_active": data.is_active}


@api_router.delete("/tenants/{tenant_id}")
async def delete_tenant(tenant_id: str, user: dict = Depends(require_roles(Role.SUPER_ADMIN))):
    """Elimina un hotel con TODOS sus datos. No hay vuelta atras.

    Mismo enfoque que CargoXprez: en vez de mantener a mano una lista ordenada
    de tablas -que se desincroniza en cuanto alguien agrega una-, se recorren
    todas las que tienen tenant_id y se reintentan las que fallan por clave
    foranea. En cada vuelta caen las hojas y en la siguiente sus padres. Si una
    vuelta entera no avanza, hay un ciclo y se aborta sin dejar el hotel a
    medias (todo va en una transaccion).

    La auditoria se escribe ANTES de borrar y en la propia bitacora del hotel
    -audit_logs.tenant_id es NOT NULL con cascade-, asi que desaparece con el.
    Queda ademas en el log del servidor, que es lo que sobrevive.
    """
    tid = id_valido(tenant_id, "tenant_id")
    async with db_pg.tx_global("eliminar un hotel: hay que barrer todas sus tablas") as conn:
        hotel = await db_pg.uno(conn, "select id, name, ruc from tenants where id = $1", tid)
        if not hotel:
            raise HTTPException(status_code=404, detail="Hotel no encontrado")

        await create_audit_log(conn, tid, user["user_id"], "tenant", "DELETE",
                               {"name": hotel["name"], "ruc": hotel["ruc"]}, None)
        logger.warning("SUPER_ADMIN %s elimina el hotel %s (%s, RUC %s) con todos sus datos",
                       user.get("email"), tid, hotel["name"], hotel["ruc"])

        pendientes = [r["t"] for r in await conn.fetch(
            "select table_name as t from information_schema.columns "
            "where table_schema = 'public' and column_name = 'tenant_id'"
        )]
        while pendientes:
            quedan = []
            for tabla in pendientes:
                try:
                    async with conn.transaction():  # savepoint: el fallo no aborta todo
                        await conn.execute(f'delete from "{tabla}" where tenant_id = $1', tid)
                except db_pg.ForeignKeyViolationError:
                    quedan.append(tabla)
            if len(quedan) == len(pendientes):
                raise HTTPException(
                    status_code=500,
                    detail="No se pudo eliminar el hotel: dependencias sin resolver en " + ", ".join(quedan),
                )
            pendientes = quedan

        await conn.execute("delete from tenants where id = $1", tid)

    return {"message": "Hotel y todos sus datos eliminados"}


@api_router.get("/tenants/{tenant_id}/stats")
async def get_tenant_stats(tenant_id: str, user: dict = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN))):
    if user["role"] != Role.SUPER_ADMIN.value and user.get("tenant_id") != tenant_id:
        raise HTTPException(status_code=403, detail="Acceso denegado")
    tid = id_valido(tenant_id, "tenant_id")
    async with db_pg.tx(user) as conn:
        if not await db_pg.valor(conn, "select 1 from tenants where id = $1", tid):
            raise HTTPException(status_code=404, detail="Hotel no encontrado")
        fila = await db_pg.uno(
            conn,
            """select
                 (select count(*) from rooms where tenant_id = $1 and is_active) as habitaciones,
                 (select count(*) from reservations where tenant_id = $1
                     and created_at >= date_trunc('month', now())) as reservas_mes,
                 (select count(*) from reservations where tenant_id = $1
                     and status = 'CHECKED_IN') as estancias_activas,
                 (select count(*) from guests where tenant_id = $1) as huespedes,
                 (select count(*) from users where tenant_id = $1) as usuarios,
                 (select count(*) from users where tenant_id = $1 and is_active) as usuarios_activos,
                 greatest(
                   (select max(created_at) from audit_logs where tenant_id = $1),
                   (select max(created_at) from reservations where tenant_id = $1)
                 ) as ultima_actividad""",
            tid,
        )
    return fila


@api_router.post("/tenants/{tenant_id}/entrar")
async def entrar_en_hotel(tenant_id: str, user: dict = Depends(require_roles(Role.SUPER_ADMIN))):
    """El SUPER_ADMIN entra a un hotel para dar soporte.

    Devuelve un token con el tenant del hotel y rol ADMIN efectivo: dentro se
    comporta como un administrador mas, RLS incluido, y todo lo que haga queda
    en los datos y la bitacora de ESE hotel. La marca en_otro_hotel es lo que
    permite volver con POST /auth/salir-de-hotel.
    """
    tid = id_valido(tenant_id, "tenant_id")
    async with db_pg.tx(user) as conn:
        hotel = await db_pg.uno(conn, "select id, name, nombre_comercial, is_active from tenants where id = $1", tid)
        if not hotel:
            raise HTTPException(status_code=404, detail="Hotel no encontrado")
        # Que un SUPER_ADMIN entre en los datos de un cliente deja rastro, y
        # el rastro va en la bitacora del hotel visitado.
        await create_audit_log(conn, tid, user["user_id"], "tenant", "SUPERADMIN_ENTER",
                               None, {"superadmin": user.get("email")})

    nombre = hotel.get("nombre_comercial") or hotel["name"]
    token = create_token(
        user["user_id"], user["email"], Role.ADMIN.value, str(hotel["id"]),
        extra={"en_otro_hotel": True, "hotel_nombre": nombre},
    )
    return {"access_token": token, "token_type": "bearer", "hotel_nombre": nombre,
            "tenant_id": str(hotel["id"]), "message": f"Dentro de {nombre}"}

class TenantSuscripcion(BaseModel):
    """Plan y estado puestos A MANO por el SUPER_ADMIN.

    Existe porque no todo el mundo paga por la pasarela: hay hoteles que pagan
    en efectivo, por Yape o por transferencia, y otros a los que se les regala
    un periodo. Sin esto, el unico camino para activar un plan era la pasarela,
    y el dueno del producto no tenia forma de honrar un pago que ya recibio.
    """
    plan_codigo: str
    subscription_status: Literal["prueba", "activa", "vencida", "suspendida", "cancelada"]
    # Hasta cuando esta pagado (o hasta cuando dura la prueba). Se guarda en
    # tenants.subscription_ends_at, la MISMA columna que escribe el IPN de
    # Izipay: las dos vias de cobro -- la pasarela y el efectivo/Yape que
    # registra el dueno a mano -- tienen que dejar el hotel en el mismo estado,
    # o "hasta cuando ha pagado este hotel" dependeria de por donde pago.
    vence: Optional[date] = None
    nota: Optional[str] = Field(default=None, max_length=300)


@api_router.put("/tenants/{tenant_id}/suscripcion")
async def update_tenant_suscripcion(
    tenant_id: str, data: TenantSuscripcion,
    user: dict = Depends(require_roles(Role.SUPER_ADMIN)),
):
    tid = id_valido(tenant_id, "tenant_id")
    async with db_pg.tx(user) as conn:
        if not await db_pg.uno(conn, "select 1 from planes where codigo = $1 and activo", data.plan_codigo):
            raise HTTPException(status_code=400, detail="Ese plan no existe o no esta activo")
        antes = await db_pg.uno(
            conn,
            "select plan_codigo, subscription_status, trial_ends_at, "
            "subscription_ends_at from tenants where id = $1",
            tid,
        )
        if not antes:
            raise HTTPException(status_code=404, detail="Hotel no encontrado")

        manual = {
            "plan_codigo": data.plan_codigo,
            "estado": data.subscription_status,
            "vence": data.vence.isoformat() if data.vence else None,
            "nota": (data.nota or "").strip() or None,
            "por": user.get("email"),
            "en": datetime.now(timezone.utc).isoformat(),
        }
        # `vence` alimenta DOS columnas y no una:
        #
        #   trial_ends_at         solo cuando el estado es 'prueba'
        #   subscription_ends_at  siempre que se indique, porque es la columna
        #                         que consulta el sistema para saber hasta
        #                         cuando esta pagado -- la misma que escribe el
        #                         IPN de Izipay.
        #
        # Sin la segunda, un hotel al que el dueno le activa el plan a mano
        # (efectivo, Yape, transferencia) quedaba 'activa' con
        # subscription_ends_at en NULL, es decir: indistinguible de uno que
        # nunca pago. La nota en settings.suscripcion_manual se conserva igual,
        # pero es una nota, no un dato comparable.
        await conn.execute(
            """update tenants
                  set plan_codigo = $2,
                      subscription_status = $3,
                      trial_ends_at = case when $3 = 'prueba' and $4::date is not null
                                           then ($4::date + interval '23 hours 59 minutes')
                                           else trial_ends_at end,
                      subscription_ends_at = case when $4::date is not null
                                                  then ($4::date + interval '23 hours 59 minutes')
                                                  else subscription_ends_at end,
                      settings = coalesce(settings, '{}'::jsonb)
                                 || jsonb_build_object('suscripcion_manual', $5::jsonb)
                where id = $1""",
            tid, data.plan_codigo, data.subscription_status, data.vence, json.dumps(manual),
        )
        # El token trae user_id, no id: con ["id"] esto levantaba KeyError y
        # el endpoint respondia 500 despues de haber hecho el UPDATE (que la
        # transaccion revertia). El plan a mano nunca llegaba a guardarse.
        await create_audit_log(
            conn, tid, user["user_id"], "tenant_suscripcion", "update",
            before={"plan_codigo": antes["plan_codigo"],
                    "subscription_status": antes["subscription_status"],
                    "subscription_ends_at": antes["subscription_ends_at"]},
            after={"plan_codigo": data.plan_codigo, "subscription_status": data.subscription_status,
                   "vence": manual["vence"], "nota": manual["nota"]},
        )
    return {"message": "Suscripcion actualizada", "suscripcion_manual": manual}


@api_router.get("/tenants/{tenant_id}/pagos")
async def listar_pagos_tenant(
    tenant_id: str, user: dict = Depends(require_roles(Role.SUPER_ADMIN)),
):
    """Historial de pagos de un hotel, para el detalle de la consola.

    Solo SUPER_ADMIN: son los cobros que el dueno del producto le hizo a ese
    hotel, no datos del hotel. Se limita a 50 -- una suscripcion mensual tarda
    cuatro anos en llegar ahi -- porque esto se pinta dentro de un dialogo y
    nadie va a leer mas.
    """
    tid = id_valido(tenant_id, "tenant_id")
    async with db_pg.tx(user) as conn:
        return await db_pg.varias(
            conn,
            """select id, plan_codigo, periodo, monto, moneda, estado, metodo,
                      izipay_order_number, izipay_transaction_id,
                      created_at, confirmado_en
                 from pagos_suscripcion
                where tenant_id = $1
                order by created_at desc
                limit 50""",
            tid,
        )


@api_router.put("/tenants/{tenant_id}/invoicing")
async def update_tenant_invoicing(tenant_id: str, config: TenantInvoicingConfig, user: dict = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN))):
    if user["role"] != Role.SUPER_ADMIN.value and user.get("tenant_id") != tenant_id:
        raise HTTPException(status_code=403, detail="Acceso denegado")

    tid = id_valido(tenant_id, "tenant_id")
    async with db_pg.tx(user) as conn:
        # Los campos van a columnas propias, no a un blob JSON: el correlativo
        # de un comprobante SUNAT tiene que poder incrementarse de forma atomica
        # y participar del unique (tenant_id, type, series, number).
        # coalesce en cada campo: lo que el cliente no manda se queda como
        # estaba. Es lo que impide que guardar el token de NubeFact reinicie
        # los correlativos -- ver la nota en TenantInvoicingConfig.
        actualizado = await conn.fetchval(
            """update tenants set
                   nubefact_ruta       = coalesce($2, nubefact_ruta),
                   nubefact_token      = coalesce($3, nubefact_token),
                   invoicing_mode      = coalesce($4, invoicing_mode),
                   boleta_series       = coalesce($5, boleta_series),
                   boleta_correlative  = coalesce($6, boleta_correlative),
                   factura_series      = coalesce($7, factura_series),
                   factura_correlative = coalesce($8, factura_correlative),
                   igv_rate            = coalesce($9, igv_rate)
               where id = $1
               returning id""",
            tid, config.nubefact_ruta, config.nubefact_token, config.invoicing_mode,
            config.boleta_series, config.boleta_correlative,
            config.factura_series, config.factura_correlative, config.igv_rate,
        )
        if not actualizado:
            raise HTTPException(status_code=404, detail="Hotel no encontrado")

        await create_audit_log(conn, tid, user["user_id"], "tenant",
                               "UPDATE_INVOICING", None, config.model_dump())
    return {"message": "Configuración de facturación actualizada"}

# ============== USER ENDPOINTS ==============
@api_router.post("/users")
async def create_user(data: UserCreate, user: dict = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN))):
    if user["role"] == Role.ADMIN.value:
        data.tenant_id = user["tenant_id"]
        if data.role == Role.SUPER_ADMIN:
            raise HTTPException(status_code=403, detail="No puede crear Super Admins")
    if len(data.password) < 8:
        raise HTTPException(status_code=400, detail="La contraseña necesita al menos 8 caracteres")

    # La coherencia rol/hotel la exige tambien la tabla (users_tenant_coherente),
    # pero un 400 con explicacion vale mas que un 500 de constraint.
    if data.role == Role.SUPER_ADMIN:
        data.tenant_id = None
    elif not data.tenant_id:
        raise HTTPException(status_code=400, detail="Indica el hotel al que pertenece el usuario")

    async with db_pg.tx_global("alta de usuario: el email es unico entre todos los hoteles") as conn:
        if await conn.fetchval("select 1 from users where email = $1", data.email):
            raise HTTPException(status_code=400, detail="Email ya registrado")
        tid = db_pg.a_uuid(data.tenant_id, "tenant_id") if data.tenant_id else None
        if tid and not await conn.fetchval("select 1 from tenants where id = $1", tid):
            raise HTTPException(status_code=404, detail="Hotel no encontrado")

        nuevo_id = await conn.fetchval(
            """insert into users (tenant_id, email, password_hash, full_name, role)
               values ($1, $2, $3, $4, $5::user_role)
               returning id""",
            tid, data.email,
            hash_password(data.password), data.full_name, data.role.value,
        )
        if tid:
            await create_audit_log(conn, tid, user["user_id"], "user", "CREATE", None,
                                   {"user_id": str(nuevo_id), "email": data.email, "role": data.role.value})
    return {"id": str(nuevo_id), "message": "Usuario creado exitosamente"}


async def _es_ultimo_admin(conn, objetivo: dict) -> bool:
    """True si `objetivo` es el unico ADMIN activo de su hotel.

    Borrarlo o desactivarlo dejaria al hotel sin nadie que pueda crear
    usuarios, y la unica salida seria que el SUPER_ADMIN entrara a mano.
    """
    if objetivo.get("role") != Role.ADMIN.value or not objetivo.get("tenant_id"):
        return False
    otros = await conn.fetchval(
        "select count(*) from users where tenant_id = $1 and role = 'ADMIN' and is_active and id <> $2",
        db_pg.a_uuid(objetivo["tenant_id"], "tenant_id"), db_pg.a_uuid(objetivo["id"], "id"),
    )
    return otros == 0


@api_router.get("/users")
async def list_users(
    tenant_id: Optional[str] = None,
    user: dict = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
):
    tid = tenant_de(user)
    # El SUPER_ADMIN puede pedir los usuarios de UN hotel (consola de Hoteles).
    # Para un ADMIN el parametro se ignora: su hotel es el del token, siempre.
    if tid is None and tenant_id:
        tid = id_valido(tenant_id, "tenant_id")
    async with db_pg.tx(user) as conn:
        # Nunca se selecciona password_hash: era la proyeccion {"password_hash": 0}
        # de Mongo, y aca se consigue simplemente no nombrando la columna.
        return await db_pg.varias(
            conn,
            """select id, tenant_id, email, full_name, role, is_active, created_at
               from users
               where ($1::uuid is null or tenant_id = $1)
               order by full_name""",
            tid,
        )

@api_router.get("/users/{user_id}", dependencies=[Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN))])
async def get_user(user_id: str, current_user: dict = Depends(get_current_user)):
    async with db_pg.tx(current_user) as conn:
        encontrado = await db_pg.uno(
            conn,
            """select id, tenant_id, email, full_name, role, is_active, created_at
               from users where id = $1""",
            id_valido(user_id, "user_id"),
        )
    if not encontrado:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if current_user['role'] != Role.SUPER_ADMIN.value and str(encontrado.get('tenant_id')) != str(current_user.get('tenant_id')):
        raise HTTPException(status_code=403, detail="No autorizado")
    return encontrado

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, data: dict = Body(...), user: dict = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN))):
    uid = id_valido(user_id, "user_id")

    # Lista blanca de columnas actualizables. Ademas de limitar que puede tocar
    # el cliente, aca cumple una segunda funcion: como los nombres se
    # interpolan en el SQL de abajo, solo pueden ser uno de estos cuatro
    # literales y nunca algo que venga del cuerpo del request.
    campos_permitidos = {"is_active", "full_name", "role", "email"}
    cambios = {k: v for k, v in data.items() if k in campos_permitidos}
    if not cambios:
        raise HTTPException(status_code=400, detail="No hay campos válidos para actualizar")

    if "role" in cambios and cambios["role"] not in [r.value for r in Role]:
        raise HTTPException(status_code=400, detail="Rol no válido")
    if "role" in cambios and cambios["role"] == Role.SUPER_ADMIN.value:
        # Un SUPER_ADMIN no tiene hotel (users_tenant_coherente); convertir a
        # alguien de un hotel en SUPER_ADMIN por aqui rompe la tabla. Se crea
        # aparte, con POST /users y tenant_id nulo.
        raise HTTPException(status_code=400, detail="El rol Super Admin no se asigna desde aquí")

    es_uno_mismo = str(user_id) == str(user["user_id"])
    if es_uno_mismo and (cambios.get("is_active") is False or "role" in cambios):
        raise HTTPException(status_code=400, detail="No puedes desactivarte ni cambiar tu propio rol")

    async with db_pg.tx_global("editar usuario: el email es unico entre todos los hoteles") as conn:
        objetivo = await db_pg.uno(conn, "select id, tenant_id, role, is_active from users where id = $1", uid)
        if not objetivo:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")

        if user["role"] != Role.SUPER_ADMIN.value and str(objetivo.get("tenant_id")) != str(user.get("tenant_id")):
            raise HTTPException(status_code=403, detail="Sin permisos para modificar este usuario")
        if objetivo.get("role") == Role.SUPER_ADMIN.value and user["role"] != Role.SUPER_ADMIN.value:
            raise HTTPException(status_code=403, detail="No autorizado para modificar a un Super Admin")

        # Que el hotel no se quede sin administrador: ni desactivando al
        # ultimo ni bajandolo de rol.
        pierde_admin = cambios.get("is_active") is False or (
            "role" in cambios and cambios["role"] != Role.ADMIN.value
        )
        if pierde_admin and objetivo.get("is_active") and await _es_ultimo_admin(conn, objetivo):
            raise HTTPException(
                status_code=400,
                detail="Es el único administrador activo del hotel. Nombra otro antes.",
            )

        if "email" in cambios and await conn.fetchval(
            "select 1 from users where email = $1 and id <> $2", cambios["email"], uid
        ):
            raise HTTPException(status_code=400, detail="Email ya registrado por otro usuario")

        # SET dinamico con parametros numerados: los nombres de columna salen de
        # la lista blanca de arriba y los valores viajan como parametros, asi
        # que no hay forma de inyectar SQL desde el cuerpo del request.
        asignaciones, valores = [], []
        for i, (col, val) in enumerate(cambios.items(), start=2):
            asignaciones.append(f"{col} = ${i}" + ("::user_role" if col == "role" else ""))
            valores.append(val)

        await conn.execute(
            f"update users set {', '.join(asignaciones)} where id = $1", uid, *valores
        )
        await create_audit_log(conn, objetivo.get("tenant_id"), user["user_id"], "user",
                               "UPDATE", None, {"user_id": user_id, **cambios})
    return {"message": "Usuario actualizado"}

@api_router.put("/users/{user_id}/password", dependencies=[Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN))])
async def reset_user_password(user_id: str, body: dict = Body(...), current_user: dict = Depends(get_current_user)):
    """Reset a user's password (admin only)"""
    new_password = body.get('password')
    if not new_password or len(new_password) < 8:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 8 caracteres")

    uid = id_valido(user_id, "user_id")
    async with db_pg.tx(current_user) as conn:
        objetivo = await db_pg.uno(conn, "select id, tenant_id, role from users where id = $1", uid)
        if not objetivo:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")

        if current_user['role'] != Role.SUPER_ADMIN.value and str(objetivo.get('tenant_id')) != str(current_user.get('tenant_id')):
            raise HTTPException(status_code=403, detail="No autorizado")
        # Un ADMIN no le reescribe la clave a un SUPER_ADMIN (seria quedarse
        # con la cuenta de la plataforma). Con RLS la fila ni se ve, pero que
        # lo diga el codigo.
        if objetivo.get('role') == Role.SUPER_ADMIN.value and current_user['role'] != Role.SUPER_ADMIN.value:
            raise HTTPException(status_code=403, detail="No autorizado")

        await conn.execute("update users set password_hash = $2 where id = $1",
                           uid, hash_password(new_password))
        # La contrasena nueva no se registra ni ofuscada: lo unico que importa
        # auditar es que alguien la cambio y a quien.
        await create_audit_log(conn, objetivo.get("tenant_id"), current_user["user_id"], "user",
                               "USER_PASSWORD_RESET", None, {"user_id": user_id})
    return {"message": "Contraseña actualizada exitosamente"}

@api_router.delete("/users/{user_id}", dependencies=[Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN))])
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a user (admin only)"""
    uid = id_valido(user_id, "user_id")
    async with db_pg.tx(current_user) as conn:
        objetivo = await db_pg.uno(conn, "select id, tenant_id, email, role from users where id = $1", uid)
        if not objetivo:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")

        if objetivo['id'] == current_user['user_id']:
            raise HTTPException(status_code=400, detail="No puedes eliminar tu propia cuenta")

        if current_user['role'] != Role.SUPER_ADMIN.value and str(objetivo.get('tenant_id')) != str(current_user.get('tenant_id')):
            raise HTTPException(status_code=403, detail="No autorizado")

        if objetivo.get('role') == Role.SUPER_ADMIN.value and current_user['role'] != Role.SUPER_ADMIN.value:
            raise HTTPException(status_code=403, detail="No autorizado para eliminar Super Administradores")

        if await _es_ultimo_admin(conn, objetivo):
            raise HTTPException(
                status_code=400,
                detail="Es el único administrador activo del hotel. Nombra otro antes de eliminarlo.",
            )

        # La auditoria se escribe ANTES del borrado: las tablas que referencian
        # al usuario (created_by, opened_by...) tienen FK contra users, asi que
        # si el usuario ya opero, el delete falla con violacion de clave
        # foranea y se responde 409 en vez de un 500 opaco.
        await create_audit_log(conn, objetivo.get("tenant_id"), current_user["user_id"], "user",
                               "USER_DELETED", None, {"user_id": user_id, "email": objetivo.get('email')})
        try:
            await conn.execute("delete from users where id = $1", uid)
        except db_pg.ForeignKeyViolationError:
            raise HTTPException(
                status_code=409,
                detail="El usuario tiene movimientos registrados y no se puede eliminar. "
                       "Desactívalo en su lugar."
            )
    return {"message": "Usuario eliminado exitosamente"}

# ============== ROOM TYPE ENDPOINTS ==============
@api_router.post("/room-types")
async def create_room_type(data: RoomTypeCreate, user: dict = Depends(require_roles(Role.ADMIN))):
    async with db_pg.tx(user) as conn:
        # amenities es text[] nativo, no JSON: es una lista homogenea de
        # etiquetas y asi se puede consultar con operadores de array.
        nuevo_id = await conn.fetchval(
            """insert into room_types (tenant_id, name, capacity, amenities, base_price)
               values ($1, $2, $3, $4, $5)
               returning id""",
            db_pg.a_uuid(user["tenant_id"]), data.name, data.capacity,
            data.amenities, data.base_price,
        )
    return {"id": str(nuevo_id), "message": "Tipo de habitación creado"}

@api_router.get("/room-types")
async def list_room_types(user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        return await db_pg.varias(
            conn,
            """select * from room_types
               where ($1::uuid is null or tenant_id = $1) and is_active
               order by name""",
            tid,
        )

@api_router.put("/room-types/{room_type_id}")
async def update_room_type(room_type_id: str, data: RoomTypeCreate, user: dict = Depends(require_roles(Role.ADMIN))):
    async with db_pg.tx(user) as conn:
        actualizado = await conn.fetchval(
            """update room_types
               set name = $3, capacity = $4, amenities = $5, base_price = $6
               where id = $1 and tenant_id = $2
               returning id""",
            id_valido(room_type_id, "room_type_id"), db_pg.a_uuid(user["tenant_id"]),
            data.name, data.capacity, data.amenities, data.base_price,
        )
    if not actualizado:
        raise HTTPException(status_code=404, detail="Tipo de habitación no encontrado")
    return {"message": "Tipo de habitación actualizado"}

# ============== RATE MANAGEMENT ENDPOINTS ==============
@api_router.post("/rates")
async def create_rate(data: RateCreate, user: dict = Depends(require_roles(Role.ADMIN))):
    async with db_pg.tx(user) as conn:
        # date_from/date_to van como `date` nativo. En Mongo se guardaban con
        # .isoformat() -- como texto --, y por eso las comparaciones de rango
        # eran comparaciones de strings que solo funcionaban de casualidad,
        # porque ISO 8601 ordena igual alfabeticamente que cronologicamente.
        nuevo_id = await conn.fetchval(
            """insert into rates (tenant_id, room_type_id, name, date_from, date_to, price, min_stay)
               values ($1, $2, $3, $4, $5, $6, $7)
               returning id""",
            db_pg.a_uuid(user["tenant_id"]), id_valido(data.room_type_id, "room_type_id"),
            data.name, data.date_from, data.date_to, data.price, data.min_stay,
        )
    return {"id": str(nuevo_id), "message": "Tarifa creada"}

@api_router.get("/rates")
async def list_rates(room_type_id: str = None, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        return await db_pg.varias(
            conn,
            """select * from rates
               where ($1::uuid is null or tenant_id = $1)
                 and is_active
                 and ($2::uuid is null or room_type_id = $2)
               order by date_from""",
            tid, db_pg.a_uuid(room_type_id, "room_type_id"),
        )

@api_router.delete("/rates/{rate_id}")
async def delete_rate(rate_id: str, user: dict = Depends(require_roles(Role.ADMIN))):
    async with db_pg.tx(user) as conn:
        borrado = await conn.fetchval(
            """update rates set is_active = false
               where id = $1 and tenant_id = $2
               returning id""",
            id_valido(rate_id, "rate_id"), db_pg.a_uuid(user["tenant_id"]),
        )
    if not borrado:
        raise HTTPException(status_code=404, detail="Tarifa no encontrada")
    return {"message": "Tarifa eliminada"}

@api_router.get("/rates/calculate")
async def calculate_rate(room_type_id: str, checkin_date: date, checkout_date: date, user: dict = Depends(get_current_user)):
    """Precio total de una estadia, noche por noche, aplicando tarifas de temporada."""
    tid = tenant_de(user)
    rtid = id_valido(room_type_id, "room_type_id")

    if checkout_date <= checkin_date:
        raise HTTPException(status_code=400, detail="La fecha de salida debe ser posterior a la de entrada")

    async with db_pg.tx(user) as conn:
        room_type = await db_pg.uno(
            conn,
            """select name, base_price from room_types
               where id = $1 and ($2::uuid is null or tenant_id = $2)""",
            rtid, tid,
        )
        if not room_type:
            raise HTTPException(status_code=404, detail="Tipo de habitación no encontrado")

        # Una sola consulta en vez de una por noche. generate_series produce las
        # noches del rango (sin incluir la de salida, que no se cobra) y el
        # LATERAL busca para cada una la tarifa de temporada vigente; si no hay,
        # cae al precio base del tipo de habitacion.
        #
        # Antes esto era un bucle en Python con un find_one por noche: una
        # estadia de dos semanas eran 14 viajes a la base solo para cotizar.
        noches = await db_pg.varias(
            conn,
            """select
                   n::date                                    as date,
                   coalesce(r.price, $4)                      as price,
                   coalesce(r.name, case when r.id is null
                                         then 'Tarifa Base'
                                         else 'Tarifa Especial' end) as rate_name
               from generate_series($1::date, $2::date - 1, interval '1 day') as n
               left join lateral (
                   select id, price, name
                   from rates
                   where room_type_id = $3
                     and is_active
                     and n::date between date_from and date_to
                   order by date_from desc
                   limit 1
               ) r on true
               order by n""",
            checkin_date, checkout_date, rtid, room_type["base_price"],
        )

    return {
        "room_type": room_type.get("name"),
        "checkin_date": checkin_date.isoformat(),
        "checkout_date": checkout_date.isoformat(),
        "nights": len(noches),
        "total": round(sum(n["price"] for n in noches), 2),
        "breakdown": noches
    }

# ============== ROOM ENDPOINTS ==============
@api_router.post("/rooms")
async def create_room(data: RoomCreate, user: dict = Depends(require_roles(Role.ADMIN))):
    async with db_pg.tx(user) as conn:
        await exigir_cupo_habitaciones(conn, user, 1)
        try:
            nuevo_id = await conn.fetchval(
                """insert into rooms (tenant_id, room_type_id, number, floor, notes)
                   values ($1, $2, $3, $4, $5)
                   returning id""",
                db_pg.a_uuid(user["tenant_id"]), id_valido(data.room_type_id, "room_type_id"),
                data.number, data.floor, data.notes,
            )
        except db_pg.UniqueViolationError:
            # Antes se consultaba primero y se insertaba despues, con lo que dos
            # altas simultaneas de la misma habitacion pasaban ambas el chequeo.
            # Ahora decide el constraint unique (tenant_id, number).
            raise HTTPException(status_code=400, detail="Número de habitación ya existe")
        except db_pg.ForeignKeyViolationError:
            raise HTTPException(status_code=400, detail="El tipo de habitación no existe")
    return {"id": str(nuevo_id), "message": "Habitación creada"}

@api_router.post("/rooms/bulk")
async def create_rooms_bulk(data: RoomBulkCreate, user: dict = Depends(require_roles(Role.ADMIN))):
    numeros = [f"{data.prefix}{data.start_number + i}" for i in range(data.count)]
    async with db_pg.tx(user) as conn:
        await exigir_cupo_habitaciones(conn, user, len(numeros))
        try:
            # unnest inserta las N habitaciones en una sola sentencia, y al ir
            # dentro de una transaccion son todas o ninguna: si la 7 de 10 choca
            # con una que ya existia, no quedan 6 habitaciones sueltas creadas.
            creadas = await conn.fetchval(
                """with nuevas as (
                       insert into rooms (tenant_id, room_type_id, number, floor)
                       select $1, $2, n, $4 from unnest($3::text[]) as n
                       returning 1
                   )
                   select count(*) from nuevas""",
                db_pg.a_uuid(user["tenant_id"]), id_valido(data.room_type_id, "room_type_id"),
                numeros, data.floor,
            )
        except db_pg.UniqueViolationError:
            raise HTTPException(
                status_code=400,
                detail="Alguno de los números de habitación del rango ya existe"
            )
        except db_pg.ForeignKeyViolationError:
            raise HTTPException(status_code=400, detail="El tipo de habitación no existe")
    return {"count": creadas, "message": f"{creadas} habitaciones creadas"}

@api_router.get("/rooms")
async def list_rooms(
    floor: Optional[int] = None,
    occupancy_status: Optional[OccupancyStatus] = None,
    housekeeping_status: Optional[HousekeepingStatus] = None,
    user: dict = Depends(get_current_user)
):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        # Un JOIN en vez de dos consultas y un diccionario armado en Python. El
        # tipo de habitacion se devuelve anidado en `room_type` con la misma
        # forma que esperaba el frontend.
        filas = await db_pg.varias(
            conn,
            """select r.*,
                      jsonb_build_object(
                          'name',       rt.name,
                          'capacity',   rt.capacity,
                          'base_price', rt.base_price
                      ) as room_type
               from rooms r
               join room_types rt on rt.id = r.room_type_id
               where ($1::uuid is null or r.tenant_id = $1)
                 and r.is_active
                 and ($2::int is null or r.floor = $2)
                 and ($3::text is null or r.occupancy_status::text = $3)
                 and ($4::text is null or r.housekeeping_status::text = $4)
               order by r.number""",
            tid, floor,
            occupancy_status.value if occupancy_status else None,
            housekeeping_status.value if housekeeping_status else None,
        )
    # base_price sale de jsonb como string (jsonb_build_object serializa el
    # numeric como numero JSON, pero asyncpg lo entrega dentro del dict ya
    # decodificado): se normaliza a float para no cambiar la forma de siempre.
    for f in filas:
        rt = f.get("room_type") or {}
        if rt.get("base_price") is not None:
            rt["base_price"] = float(rt["base_price"])
    return filas

@api_router.put("/rooms/{room_id}/status")
async def update_room_status(room_id: str, occupancy: Optional[OccupancyStatus] = None, housekeeping: Optional[HousekeepingStatus] = None, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    rid = id_valido(room_id, "room_id")

    if not occupancy and not housekeeping:
        return {"message": "Estado de habitación actualizado"}

    async with db_pg.tx(user) as conn:
        # El update devuelve los valores anteriores y los nuevos de una sola vez:
        # antes se leia la habitacion, se escribia y se volvia a usar la lectura
        # vieja para la bitacora, con lo que dos cambios simultaneos podian
        # registrar una transicion que nunca ocurrio.
        cambio = await db_pg.uno(
            conn,
            """update rooms r set
                   occupancy_status    = coalesce($3::occupancy_status,   r.occupancy_status),
                   housekeeping_status = coalesce($4::housekeeping_state, r.housekeeping_status)
               from rooms viejo
               where r.id = viejo.id
                 and r.id = $1
                 and ($2::uuid is null or r.tenant_id = $2)
               returning viejo.occupancy_status    as from_occupancy,
                         r.occupancy_status        as to_occupancy,
                         viejo.housekeeping_status as from_housekeeping,
                         r.housekeeping_status     as to_housekeeping,
                         r.tenant_id               as tenant_id""",
            rid, tid,
            occupancy.value if occupancy else None,
            housekeeping.value if housekeeping else None,
        )
        if not cambio:
            raise HTTPException(status_code=404, detail="Habitación no encontrada")

        await conn.execute(
            """insert into housekeeping_logs
                   (tenant_id, room_id, from_occupancy, to_occupancy,
                    from_housekeeping, to_housekeeping, by_user)
               values ($1, $2, $3::occupancy_status, $4::occupancy_status,
                       $5::housekeeping_state, $6::housekeeping_state, $7)""",
            db_pg.a_uuid(cambio["tenant_id"]), rid,
            cambio["from_occupancy"], cambio["to_occupancy"],
            cambio["from_housekeeping"], cambio["to_housekeeping"],
            id_valido(user["user_id"], "user_id"),
        )

        await create_audit_log(
            conn, cambio["tenant_id"], user["user_id"], "room", "STATUS_CHANGE",
            {"occupancy_status": cambio["from_occupancy"], "housekeeping_status": cambio["from_housekeeping"]},
            {"occupancy_status": cambio["to_occupancy"], "housekeeping_status": cambio["to_housekeeping"]},
        )

    return {"message": "Estado de habitación actualizado"}

# ============== GUEST ENDPOINTS ==============
@api_router.post("/guests")
async def create_guest(data: GuestCreate, user: dict = Depends(get_current_user)):
    async with db_pg.tx(user) as conn:
        # El endpoint siempre fue idempotente: si el huesped ya existe devuelve
        # el que hay en vez de fallar (recepcion vuelve a cargar el mismo DNI en
        # cada estadia). Ahora eso lo resuelve ON CONFLICT contra el unique
        # (tenant_id, doc_type, doc_number), en una sola ida a la base y sin la
        # carrera que tenia el consultar-y-luego-insertar.
        huesped = await db_pg.uno(
            conn,
            """insert into guests (tenant_id, doc_type, doc_number, full_name,
                                   phone, email, nationality, address)
               values ($1, $2::doc_type, $3, $4, $5, $6, $7, $8)
               on conflict (tenant_id, doc_type, doc_number) do update
                   set full_name = excluded.full_name
               returning *""",
            db_pg.a_uuid(user["tenant_id"]), data.doc_type.value, data.doc_number,
            data.full_name, data.phone, data.email, data.nationality, data.address,
        )
    return huesped

@api_router.get("/guests")
async def list_guests(search: Optional[str] = None, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        # El texto de busqueda va como parametro y se escapa con `like` sobre
        # columnas concretas. Antes se inyectaba crudo en un $regex de Mongo:
        # ademas de ser lento, un patron malicioso como "(a+)+$" podia colgar la
        # consulta contra toda la tabla.
        return await db_pg.varias(
            conn,
            """select * from guests
               where ($1::uuid is null or tenant_id = $1)
                 and ($2::text is null
                      or full_name  ilike '%' || $2 || '%'
                      or doc_number ilike '%' || $2 || '%'
                      or email      ilike '%' || $2 || '%')
               order by full_name""",
            tid, search,
        )

@api_router.get("/guests/{guest_id}")
async def get_guest(guest_id: str, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        huesped = await db_pg.uno(
            conn,
            """select * from guests
               where id = $1 and ($2::uuid is null or tenant_id = $2)""",
            id_valido(guest_id, "guest_id"), tid,
        )
    if not huesped:
        raise HTTPException(status_code=404, detail="Huésped no encontrado")
    return huesped

# ============== RESERVATION ENDPOINTS ==============
@api_router.post("/reservations")
async def create_reservation(data: ReservationCreate, user: dict = Depends(exigir_suscripcion)):
    tid = db_pg.a_uuid(user["tenant_id"], "tenant_id")

    if data.checkout_date <= data.checkin_date:
        raise HTTPException(status_code=400, detail="La fecha de salida debe ser posterior a la de entrada")

    async with db_pg.tx(user) as conn:
        # Serializa la creacion de reservas DE ESTE HOTEL durante la
        # transaccion. Cubre dos carreras que existian:
        #
        #  1. El codigo salia de count(*)+1. Dos recepcionistas creando reserva
        #     a la vez obtenian el mismo "RES-000042"; y al cancelar una, el
        #     contador retrocedia y el siguiente codigo repetia uno ya usado.
        #  2. El solapamiento se consultaba y despues se insertaba, que es
        #     precisamente como se produce un overbooking.
        #
        # El lock es por hotel, no de tabla, asi que no molesta a los otros
        # clientes, y se libera solo al terminar la transaccion.
        await conn.execute("select pg_advisory_xact_lock(hashtext($1))", str(tid))

        # El correlativo sale del maximo real, no de contar filas: cancelar una
        # reserva ya no hace retroceder la numeracion.
        siguiente = await conn.fetchval(
            """select coalesce(max(substring(code from 'RES-([0-9]+)$')::int), 0) + 1
               from reservations where tenant_id = $1""",
            tid,
        )
        code = f"RES-{siguiente:06d}"

        if data.room_id:
            # Dos rangos se solapan si cada uno empieza antes de que termine el
            # otro. La reserva que sale el dia que entra la siguiente NO se
            # solapa: por eso < y > y no <= y >=.
            ocupada = await conn.fetchval(
                """select 1 from reservations
                   where tenant_id = $1
                     and room_id = $2
                     and status in ('CONFIRMED', 'CHECKED_IN')
                     and checkin_date < $4
                     and checkout_date > $3
                   limit 1""",
                tid, id_valido(data.room_id, "room_id"),
                data.checkin_date, data.checkout_date,
            )
            if ocupada:
                raise HTTPException(status_code=400, detail="Habitación no disponible para esas fechas")

        try:
            nuevo_id = await conn.fetchval(
                """insert into reservations
                       (tenant_id, code, guest_id, room_type_id, room_id,
                        checkin_date, checkout_date, adults, children,
                        total_estimated, deposit_amount, deposit_status,
                        status, source, notes, created_by)
                   values ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12,
                           'CONFIRMED', $13, $14, $15)
                   returning id""",
                tid, code,
                id_valido(data.guest_id, "guest_id"),
                id_valido(data.room_type_id, "room_type_id"),
                id_valido(data.room_id, "room_id"),
                data.checkin_date, data.checkout_date,
                data.adults, data.children,
                data.total_estimated, data.deposit_amount,
                "PENDING" if data.deposit_amount > 0 else "NA",
                data.source, data.notes,
                id_valido(user["user_id"], "user_id"),
            )
        except db_pg.ForeignKeyViolationError:
            raise HTTPException(
                status_code=400,
                detail="El huésped, el tipo de habitación o la habitación indicada no existe"
            )

        await create_audit_log(conn, tid, user["user_id"], "reservation", "CREATE", None, {"code": code})

    return {"id": str(nuevo_id), "code": code, "message": "Reserva creada exitosamente"}

@api_router.get("/reservations")
async def list_reservations(
    status: Optional[ReservationStatus] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    search: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        # El huesped se trae con un JOIN, no con una segunda consulta y un
        # diccionario en Python. La forma de la respuesta no cambia: `guest`
        # sigue siendo un objeto anidado con los mismos tres campos.
        return await db_pg.varias(
            conn,
            """select r.*,
                      jsonb_build_object(
                          'full_name',  g.full_name,
                          'doc_type',   g.doc_type,
                          'doc_number', g.doc_number
                      ) as guest
               from reservations r
               join guests g on g.id = r.guest_id
               where ($1::uuid is null or r.tenant_id = $1)
                 and ($2::text is null or r.status::text = $2)
                 and ($3::date is null or r.checkin_date  >= $3)
                 and ($4::date is null or r.checkout_date <= $4)
                 and ($5::text is null or r.code ilike '%' || $5 || '%')
               order by r.checkin_date desc""",
            tid,
            status.value if status else None,
            from_date, to_date, search,
        )

# ============== GROUP RESERVATIONS (must be before {reservation_id} routes) ==============
@api_router.post("/reservations/group")
async def create_group_reservation(
    data: GroupReservationCreate,
    user: dict = Depends(get_current_user)
):
    """Crea una reserva de grupo: N habitaciones bloqueadas a nombre del grupo."""
    tid = db_pg.a_uuid(user["tenant_id"], "tenant_id")

    if data.checkin_date >= data.checkout_date:
        raise HTTPException(status_code=400, detail="Fecha de checkout debe ser posterior a checkin")

    nights = (data.checkout_date - data.checkin_date).days
    creador = id_valido(user["user_id"], "user_id")

    async with db_pg.tx(user) as conn:
        # Mismo lock por hotel que en el alta individual: los correlativos GRP-
        # y RES- se generan aca y no pueden repetirse. Antes ambos salian de
        # count(*)+1, y el de las reservas se recalculaba DENTRO del bucle: una
        # consulta por habitacion, y con dos grupos creandose a la vez los
        # codigos se pisaban.
        await conn.execute("select pg_advisory_xact_lock(hashtext($1))", str(tid))

        n_grupo = await conn.fetchval(
            """select coalesce(max(substring(code from 'GRP-([0-9]+)$')::int), 0) + 1
               from group_reservations where tenant_id = $1""",
            tid,
        )
        group_code = f"GRP-{n_grupo:06d}"

        # Los tipos de habitacion pedidos, resueltos de una vez. Antes era un
        # find_one por linea del pedido.
        pedidos = [(id_valido(r.get("room_type_id"), "room_type_id"), int(r.get("quantity", 1)))
                   for r in data.rooms]
        if not pedidos:
            raise HTTPException(status_code=400, detail="La reserva grupal no incluye habitaciones")

        precios = {
            f["id"]: f["base_price"]
            for f in await db_pg.varias(
                conn,
                "select id, base_price from room_types where tenant_id = $1 and id = any($2::uuid[])",
                tid, [p[0] for p in pedidos],
            )
        }
        for rtid, _ in pedidos:
            if str(rtid) not in precios:
                raise HTTPException(status_code=400, detail=f"Tipo de habitación no encontrado: {rtid}")

        total_rooms = sum(q for _, q in pedidos)
        total_estimated = sum(nights * precios[str(rtid)] * q for rtid, q in pedidos)

        group_id = await conn.fetchval(
            """insert into group_reservations
                   (tenant_id, code, group_name, contact_name, contact_phone,
                    contact_email, checkin_date, checkout_date, nights, adults,
                    children, total_rooms, total_estimated, deposit_amount,
                    status, notes, created_by)
               values ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,'CONFIRMED',$15,$16)
               returning id""",
            tid, group_code, data.group_name, data.contact_name, data.contact_phone,
            data.contact_email, data.checkin_date, data.checkout_date, nights,
            data.adults, data.children, total_rooms, total_estimated,
            data.deposit_amount, data.notes, creador,
        )

        n_reserva = await conn.fetchval(
            """select coalesce(max(substring(code from 'RES-([0-9]+)$')::int), 0)
               from reservations where tenant_id = $1""",
            tid,
        )

        # Las N reservas del grupo en una sola sentencia. generate_series expande
        # la cantidad pedida de cada tipo, y el correlativo sale de row_number()
        # sumado al ultimo usado -- sin volver a consultar por cada habitacion.
        # Van sin guest_id (se asigna en el check-in) y sin room_id (recepcion
        # ubica al grupo al llegar), que es justo lo que permite el CHECK
        # reservations_huesped_o_grupo.
        ids = await conn.fetch(
            """with pedido as (
                   select unnest($1::uuid[]) as room_type_id,
                          unnest($2::int[])  as cantidad,
                          unnest($3::numeric[]) as precio
               ),
               expandido as (
                   select p.room_type_id,
                          p.precio * $4 as total_linea,
                          p.cantidad,
                          row_number() over () as n
                   from pedido p, generate_series(1, p.cantidad)
               )
               insert into reservations
                   (tenant_id, code, group_id, room_type_id, checkin_date,
                    checkout_date, adults, children, total_estimated,
                    source, status, notes, created_by)
               select $5,
                      'RES-' || lpad(($6 + e.n)::text, 6, '0'),
                      $7, e.room_type_id, $8, $9,
                      greatest(1, $10 / e.cantidad), 0, e.total_linea,
                      'GRUPO', 'CONFIRMED', $11, $12
               from expandido e
               returning id""",
            [p[0] for p in pedidos], [p[1] for p in pedidos],
            [precios[str(p[0])] for p in pedidos], nights,
            tid, n_reserva, group_id, data.checkin_date, data.checkout_date,
            data.adults, f"Grupo: {data.group_name}", creador,
        )

        await create_audit_log(conn, tid, user["user_id"], "group_reservation", "CREATE",
                               None, {"code": group_code, "rooms": len(ids)})

    return {
        "id": str(group_id),
        "code": group_code,
        "reservations_created": len(ids),
        "total_estimated": float(total_estimated),
        "message": "Reserva grupal creada exitosamente"
    }

@api_router.get("/reservations/groups")
async def list_group_reservations(
    status: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Lista las reservas de grupo."""
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        return await db_pg.varias(
            conn,
            """select * from group_reservations
               where ($1::uuid is null or tenant_id = $1)
                 and ($2::text is null or status = $2)
               order by created_at desc""",
            tid, status,
        )

@api_router.get("/reservations/groups/{group_id}")
async def get_group_reservation(group_id: str, user: dict = Depends(get_current_user)):
    """Una reserva de grupo con el detalle de sus habitaciones."""
    tid = tenant_de(user)
    gid = id_valido(group_id, "group_id")
    async with db_pg.tx(user) as conn:
        grupo = await db_pg.uno(
            conn,
            """select * from group_reservations
               where id = $1 and ($2::uuid is null or tenant_id = $2)""",
            gid, tid,
        )
        if not grupo:
            raise HTTPException(status_code=404, detail="Reserva grupal no encontrada")

        # Las reservas del grupo salen por la FK group_id, no por comparar el
        # texto del codigo como antes: si alguien renombrara un codigo, el
        # vinculo se perdia en silencio.
        grupo["reservation_details"] = await db_pg.varias(
            conn, "select * from reservations where group_id = $1 order by code", gid
        )
    return grupo

@api_router.get("/reservations/{reservation_id}")
async def get_reservation(reservation_id: str, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        # Tres consultas se vuelven una. Los LEFT JOIN son necesarios y no
        # simples JOIN: una reserva de grupo puede no tener huesped todavia, y
        # una reserva sin ubicar no tiene habitacion. to_jsonb(...) devuelve
        # NULL en esos casos, que es exactamente lo que devolvia antes.
        reserva = await db_pg.uno(
            conn,
            """select r.*,
                      to_jsonb(g.*) as guest,
                      to_jsonb(h.*) as room
               from reservations r
               left join guests g on g.id = r.guest_id
               left join rooms  h on h.id = r.room_id
               where r.id = $1 and ($2::uuid is null or r.tenant_id = $2)""",
            id_valido(reservation_id, "reservation_id"), tid,
        )
    if not reserva:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    return reserva

@api_router.put("/reservations/{reservation_id}")
async def update_reservation(reservation_id: str, data: ReservationUpdate, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    rid = id_valido(reservation_id, "reservation_id")

    if data.status in (ReservationStatus.CANCELLED, ReservationStatus.NO_SHOW) and not data.cancel_reason:
        raise HTTPException(status_code=400, detail="Se requiere motivo para cancelar/no-show")

    async with db_pg.tx(user) as conn:
        antes = await db_pg.uno(
            conn,
            """select * from reservations
               where id = $1 and ($2::uuid is null or tenant_id = $2)
               for update""",
            rid, tid,
        )
        if not antes:
            raise HTTPException(status_code=404, detail="Reserva no encontrada")

        try:
            # coalesce deja pasar solo los campos que vinieron: es el
            # equivalente del $set parcial, en una sola sentencia.
            despues = await db_pg.uno(
                conn,
                """update reservations set
                       checkin_date  = coalesce($3::date, checkin_date),
                       checkout_date = coalesce($4::date, checkout_date),
                       room_id       = coalesce($5::uuid, room_id),
                       notes         = coalesce($6::text, notes),
                       status        = coalesce($7::reservation_status, status),
                       cancel_reason = coalesce($8::text, cancel_reason)
                   where id = $1 and ($2::uuid is null or tenant_id = $2)
                   returning *""",
                rid, tid, data.checkin_date, data.checkout_date,
                id_valido(data.room_id, "room_id"), data.notes,
                data.status.value if data.status else None,
                data.cancel_reason,
            )
        except db_pg.ExclusionViolationError:
            # Cambiar fechas o habitacion puede chocar con otra reserva. Lo
            # detecta el constraint de db/migrations/001, no la aplicacion.
            raise HTTPException(
                status_code=409,
                detail="La habitación ya está reservada para esas fechas"
            )
        except db_pg.CheckViolationError:
            raise HTTPException(
                status_code=400,
                detail="La fecha de salida debe ser posterior a la de entrada"
            )

        await create_audit_log(conn, antes["tenant_id"], user["user_id"],
                               "reservation", "UPDATE", antes, despues)

    return {"message": "Reserva actualizada"}

@api_router.post("/reservations/{reservation_id}/assign-room")
async def assign_room(reservation_id: str, room_id: str = Body(..., embed=True), user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    rid = id_valido(reservation_id, "reservation_id")
    hid = id_valido(room_id, "room_id")

    async with db_pg.tx(user) as conn:
        habitacion = await db_pg.uno(
            conn,
            """select occupancy_status, housekeeping_status from rooms
               where id = $1 and ($2::uuid is null or tenant_id = $2)""",
            hid, tid,
        )
        if not habitacion:
            raise HTTPException(status_code=404, detail="Habitación no encontrada")
        if habitacion["housekeeping_status"] == HousekeepingStatus.OUT_OF_ORDER.value:
            raise HTTPException(status_code=400, detail="Habitación fuera de servicio")
        if habitacion["occupancy_status"] != OccupancyStatus.VACANT.value:
            raise HTTPException(status_code=400, detail="Habitación ocupada")

        try:
            asignada = await conn.fetchval(
                """update reservations set room_id = $3
                   where id = $1 and ($2::uuid is null or tenant_id = $2)
                   returning id""",
                rid, tid, hid,
            )
        except db_pg.ExclusionViolationError:
            # La habitacion esta libre AHORA, pero ya tiene otra reserva que
            # pisa estas fechas. Sin el constraint, esto se descubria el dia del
            # check-in con los dos huespedes en el mostrador.
            raise HTTPException(
                status_code=409,
                detail="La habitación ya tiene otra reserva en esas fechas"
            )
        if not asignada:
            raise HTTPException(status_code=404, detail="Reserva no encontrada")

    return {"message": "Habitación asignada"}

# ============== CHECK-IN / CHECK-OUT ENDPOINTS ==============
@api_router.post("/reservations/{reservation_id}/checkin")
async def perform_checkin(reservation_id: str, user: dict = Depends(exigir_suscripcion)):
    tid = tenant_de(user)
    rid = id_valido(reservation_id, "reservation_id")

    # Las cinco escrituras del check-in van en UNA transaccion. Antes eran
    # sueltas: si fallaba a mitad quedaba una reserva marcada como ingresada sin
    # folio, o un folio huerfano, y habia que arreglarlo a mano en la base.
    async with db_pg.tx(user) as conn:
        reserva = await db_pg.uno(
            conn,
            """select r.*, h.housekeeping_status
               from reservations r
               left join rooms h on h.id = r.room_id
               where r.id = $1 and ($2::uuid is null or r.tenant_id = $2)
               for update of r""",
            rid, tid,
        )
        if not reserva:
            raise HTTPException(status_code=404, detail="Reserva no encontrada")
        if reserva["status"] != ReservationStatus.CONFIRMED.value:
            raise HTTPException(status_code=400, detail="Reserva no está confirmada")
        if not reserva.get("room_id"):
            raise HTTPException(status_code=400, detail="No hay habitación asignada")
        if reserva["housekeeping_status"] == HousekeepingStatus.OUT_OF_ORDER.value:
            raise HTTPException(status_code=400, detail="Habitación fuera de servicio")

        hotel = db_pg.a_uuid(reserva["tenant_id"])
        habitacion = db_pg.a_uuid(reserva["room_id"])

        # El unique (reservation_id) de stays impide hacer dos check-in de la
        # misma reserva aunque lleguen a la vez.
        stay_id = await conn.fetchval(
            """insert into stays (tenant_id, reservation_id, guest_id, room_id,
                                  checkin_at, status, created_by)
               values ($1, $2, $3, $4, now(), 'OPEN', $5)
               returning id""",
            hotel, rid, db_pg.a_uuid(reserva.get("guest_id")), habitacion,
            id_valido(user["user_id"], "user_id"),
        )

        # Igual con folios: unique (reservation_id), un folio por reserva.
        folio_id = await conn.fetchval(
            """insert into folios (tenant_id, reservation_id, stay_id, status)
               values ($1, $2, $3, 'OPEN')
               returning id""",
            hotel, rid, stay_id,
        )

        # stay_id y folio_id ya no se copian dentro de la reserva: la relacion
        # vive en el lado hijo, con FK de verdad, y se consulta por
        # reservation_id. Asi no hay dos sitios que puedan discrepar.
        await conn.execute(
            "update reservations set status = 'CHECKED_IN' where id = $1", rid
        )
        await conn.execute(
            "update rooms set occupancy_status = 'OCCUPIED' where id = $1", habitacion
        )

        await create_audit_log(conn, hotel, user["user_id"], "reservation", "CHECKIN",
                               None, {"reservation_id": reservation_id, "stay_id": str(stay_id)})

    return {"stay_id": str(stay_id), "folio_id": str(folio_id), "message": "Check-in realizado exitosamente"}

@api_router.post("/reservations/{reservation_id}/checkout")
async def perform_checkout(reservation_id: str, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    rid = id_valido(reservation_id, "reservation_id")

    async with db_pg.tx(user) as conn:
        # stay y folio se resuelven por FK, no leyendo campos copiados dentro de
        # la reserva.
        #
        # El FOR UPDATE va SOLO sobre `reservations`: Postgres no admite
        # bloquear el lado nullable de un LEFT JOIN ("FOR UPDATE cannot be
        # applied to the nullable side of an outer join"), y stays/folios entran
        # por LEFT JOIN porque una reserva podria no tenerlos.
        estado = await db_pg.uno(
            conn,
            """select r.tenant_id, r.status, r.room_id,
                      s.id as stay_id, f.id as folio_id
               from reservations r
               left join stays  s on s.reservation_id = r.id
               left join folios f on f.reservation_id = r.id
               where r.id = $1 and ($2::uuid is null or r.tenant_id = $2)
               for update of r""",
            rid, tid,
        )
        if not estado:
            raise HTTPException(status_code=404, detail="Reserva no encontrada")
        if estado["status"] != ReservationStatus.CHECKED_IN.value:
            raise HTTPException(status_code=400, detail="Reserva no está en check-in")

        # El folio se bloquea aparte, ya sin join. Importa hacerlo antes de leer
        # el saldo: si no, un cobro que entrara entre la lectura y el cierre se
        # perderia y el huesped se iria con deuda dada por saldada.
        saldo = None
        if estado.get("folio_id"):
            saldo = await conn.fetchval(
                "select balance from folios where id = $1 for update",
                db_pg.a_uuid(estado["folio_id"]),
            )
        if saldo is not None and saldo > 0:
            raise HTTPException(
                status_code=400,
                detail=f"Folio tiene saldo pendiente: S/ {saldo:.2f}"
            )

        hotel = db_pg.a_uuid(estado["tenant_id"])
        habitacion = db_pg.a_uuid(estado["room_id"])
        stay_id = db_pg.a_uuid(estado["stay_id"])

        await conn.execute(
            "update stays set checkout_at = now(), status = 'CLOSED' where id = $1", stay_id
        )
        await conn.execute(
            "update folios set status = 'CLOSED' where id = $1", db_pg.a_uuid(estado["folio_id"])
        )
        await conn.execute(
            "update reservations set status = 'CHECKED_OUT' where id = $1", rid
        )
        # La habitacion queda libre pero sucia: nadie puede asignarla hasta que
        # housekeeping la marque limpia.
        await conn.execute(
            """update rooms set occupancy_status = 'VACANT',
                                housekeeping_status = 'DIRTY'
               where id = $1""",
            habitacion,
        )
        await conn.execute(
            """insert into housekeeping_tasks (tenant_id, room_id, stay_id, priority, status)
               values ($1, $2, $3, 'HIGH', 'OPEN')""",
            hotel, habitacion, stay_id,
        )

        await create_audit_log(conn, hotel, user["user_id"], "reservation", "CHECKOUT",
                               None, {"reservation_id": reservation_id})

    return {"message": "Check-out realizado exitosamente"}

# ============== FOLIO ENDPOINTS ==============
@api_router.get("/folios/{folio_id}")
async def get_folio(folio_id: str, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    fid = id_valido(folio_id, "folio_id")
    async with db_pg.tx(user) as conn:
        # Cinco consultas encadenadas se vuelven una. Los subselect agregados
        # devuelven ya los arrays de cargos y pagos, y el huesped sale del stay
        # por LEFT JOIN (puede no haberlo en una reserva de grupo sin asignar).
        folio = await db_pg.uno(
            conn,
            """select f.*,
                      coalesce((select jsonb_agg(to_jsonb(c.*) order by c.created_at)
                                from charges c
                                where c.folio_id = f.id and c.status = 'ACTIVE'), '[]'::jsonb) as charges,
                      coalesce((select jsonb_agg(to_jsonb(p.*) order by p.created_at)
                                from payments p
                                where p.folio_id = f.id and p.status = 'ACTIVE'), '[]'::jsonb) as payments,
                      to_jsonb(g.*) as guest
               from folios f
               left join stays  s on s.id = f.stay_id
               left join guests g on g.id = s.guest_id
               where f.id = $1 and ($2::uuid is null or f.tenant_id = $2)""",
            fid, tid,
        )
    if not folio:
        raise HTTPException(status_code=404, detail="Folio no encontrado")
    return folio

@api_router.post("/folios/{folio_id}/charges")
async def add_charge(folio_id: str, data: ChargeCreate, user: dict = Depends(exigir_suscripcion)):
    tid = tenant_de(user)
    fid = id_valido(folio_id, "folio_id")

    async with db_pg.tx(user) as conn:
        folio = await db_pg.uno(
            conn,
            """select f.id, f.tenant_id, f.status, t.igv_rate
               from folios f
               join tenants t on t.id = f.tenant_id
               where f.id = $1 and ($2::uuid is null or f.tenant_id = $2)
               for update of f""",
            fid, tid,
        )
        if not folio:
            raise HTTPException(status_code=404, detail="Folio no encontrado")
        if folio["status"] != "OPEN":
            raise HTTPException(status_code=400, detail="Folio cerrado")

        # El IGV lo calcula Postgres en numeric, no Python en float. Con float,
        # 3 noches a 85.50 mas IGV daban un total que difiere en centimos del
        # que despues imprime la boleta; y la suma de esos centimos es la que
        # hace que el arqueo del turno no cuadre.
        #
        # La tasa sale de la columna igv_rate del hotel, que ahora existe una
        # sola vez (antes estaba duplicada dentro de invoicing_config y en la
        # raiz, y este endpoint leia la anidada).
        nuevo = await db_pg.uno(
            conn,
            """with calculo as (
                   select round($4::numeric * $5::numeric, 2) as subtotal
               ),
               importes as (
                   select subtotal,
                          case when $6 = 'IGV'
                               then round(subtotal * ($7::numeric / 100), 2)
                               else 0 end as igv
                   from calculo
               ),
               insertado as (
                   insert into charges (tenant_id, folio_id, concept, category,
                                        quantity, unit_price, subtotal, tax_type,
                                        igv_amount, total, created_by)
                   select $1, $2, $3, $8, $4, $5, subtotal, $6, igv, subtotal + igv, $9
                   from importes
                   returning id, total
               ),
               ajuste as (
                   update folios f
                   set total_charges = f.total_charges + i.total,
                       balance       = f.balance + i.total
                   from insertado i
                   where f.id = $2
                   returning 1
               )
               select id, total from insertado""",
            db_pg.a_uuid(folio["tenant_id"]), fid, data.concept,
            data.quantity, data.unit_price, data.tax_type, folio["igv_rate"],
            data.category, id_valido(user["user_id"], "user_id"),
        )

    return {"id": str(nuevo["id"]), "message": "Cargo agregado"}

@api_router.post("/folios/{folio_id}/charges/{charge_id}/void")
async def void_charge(folio_id: str, charge_id: str, data: VoidRequest, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    fid = id_valido(folio_id, "folio_id")
    cid = id_valido(charge_id, "charge_id")

    async with db_pg.tx(user) as conn:
        # El UPDATE lleva `and status = 'ACTIVE'` en el WHERE: si el cargo ya se
        # anulo (o dos anulaciones llegan a la vez), la segunda no toca nada y
        # devuelve cero filas. Antes se comprobaba antes de escribir, con lo que
        # una doble anulacion descontaba el importe dos veces del folio y
        # dejaba la cuenta del huesped en negativo.
        anulado = await db_pg.uno(
            conn,
            """update charges
               set status = 'VOIDED', void_reason = $4,
                   voided_by = $5, voided_at = now()
               where id = $1 and folio_id = $2 and status = 'ACTIVE'
                 and ($3::uuid is null or tenant_id = $3)
               returning tenant_id, total""",
            cid, fid, tid, data.reason, id_valido(user["user_id"], "user_id"),
        )
        if not anulado:
            existe = await conn.fetchval(
                "select status from charges where id = $1 and folio_id = $2", cid, fid
            )
            if existe:
                raise HTTPException(status_code=400, detail="Cargo ya anulado")
            raise HTTPException(status_code=404, detail="Cargo no encontrado")

        await conn.execute(
            """update folios
               set total_charges = total_charges - $2,
                   balance       = balance - $2
               where id = $1""",
            fid, anulado["total"],
        )

        await create_audit_log(conn, anulado["tenant_id"], user["user_id"], "charge", "VOID",
                               {"id": charge_id, "total": anulado["total"]}, {"reason": data.reason})

    return {"message": "Cargo anulado"}

@api_router.post("/folios/{folio_id}/payments")
async def add_payment(folio_id: str, data: PaymentCreate, user: dict = Depends(exigir_suscripcion)):
    tid = tenant_de(user)
    fid = id_valido(folio_id, "folio_id")

    async with db_pg.tx(user) as conn:
        folio = await db_pg.uno(
            conn,
            """select id, tenant_id from folios
               where id = $1 and ($2::uuid is null or tenant_id = $2)
               for update""",
            fid, tid,
        )
        if not folio:
            raise HTTPException(status_code=404, detail="Folio no encontrado")

        hotel = db_pg.a_uuid(folio["tenant_id"])
        turno = await conn.fetchval(
            "select id from cash_shifts where tenant_id = $1 and status = 'OPEN'", hotel
        )
        if not turno:
            raise HTTPException(status_code=400, detail="No hay caja abierta")

        # El pago y el ajuste del saldo, en una sola sentencia y una sola
        # transaccion: no puede quedar un pago cobrado que el folio no refleje.
        nuevo = await db_pg.uno(
            conn,
            """with insertado as (
                   insert into payments (tenant_id, folio_id, cash_shift_id,
                                         method, amount, reference, created_by)
                   values ($1, $2, $3, $4::payment_method, $5, $6, $7)
                   returning id, amount
               ),
               ajuste as (
                   update folios f
                   set total_payments = f.total_payments + i.amount,
                       balance        = f.balance - i.amount
                   from insertado i
                   where f.id = $2
                   returning 1
               )
               select id from insertado""",
            hotel, fid, turno, data.method.value, data.amount, data.reference,
            id_valido(user["user_id"], "user_id"),
        )

    return {"id": str(nuevo["id"]), "message": "Pago registrado"}

# ============== CASH SHIFT ENDPOINTS ==============
@api_router.post("/cash-shifts/open")
async def open_cash_shift(data: CashShiftOpen, user: dict = Depends(get_current_user)):
    tid = db_pg.a_uuid(user["tenant_id"], "tenant_id")
    async with db_pg.tx(user) as conn:
        try:
            nuevo_id = await conn.fetchval(
                """insert into cash_shifts (tenant_id, opening_amount, opened_by, status)
                   values ($1, $2, $3, 'OPEN')
                   returning id""",
                tid, data.opening_amount, id_valido(user["user_id"], "user_id"),
            )
        except db_pg.UniqueViolationError:
            # Lo decide el indice unico parcial de db/indexes.sql, no un chequeo
            # previo: dos aperturas simultaneas ya no crean dos cajas.
            raise HTTPException(status_code=400, detail="Ya hay una caja abierta")

        await create_audit_log(conn, tid, user["user_id"], "cash_shift", "OPEN",
                               None, {"opening_amount": data.opening_amount})

    return {"id": str(nuevo_id), "message": "Caja abierta"}

@api_router.get("/cash-shifts/current")
async def get_current_cash_shift(user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        # Los totales por metodo de pago salen agrupados de la propia base, en
        # lugar del pipeline de agregacion de Mongo mas un bucle en Python.
        turno = await db_pg.uno(
            conn,
            """select s.*,
                      coalesce((select jsonb_object_agg(p.method, p.suma)
                                from (select method, sum(amount) as suma
                                      from payments
                                      where cash_shift_id = s.id and status = 'ACTIVE'
                                      group by method) p), '{}'::jsonb) as totals,
                      coalesce((select sum(amount) from payments
                                where cash_shift_id = s.id and status = 'ACTIVE'), 0) as total_payments
               from cash_shifts s
               where ($1::uuid is null or s.tenant_id = $1) and s.status = 'OPEN'""",
            tid,
        )
    return turno

@api_router.post("/cash-shifts/{shift_id}/close")
async def close_cash_shift(shift_id: str, data: CashShiftClose, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    sid = id_valido(shift_id, "shift_id")

    async with db_pg.tx(user) as conn:
        turno = await db_pg.uno(
            conn,
            """select id, tenant_id, status, opening_amount from cash_shifts
               where id = $1 and ($2::uuid is null or tenant_id = $2)
               for update""",
            sid, tid,
        )
        if not turno:
            raise HTTPException(status_code=404, detail="Caja no encontrada")
        if turno["status"] != "OPEN":
            raise HTTPException(status_code=400, detail="Caja no está abierta")

        # El efectivo esperado ahora INCLUYE los movimientos de caja. Antes era
        # solo apertura + cobros en efectivo, y los cash_movements -- las
        # entradas y salidas por compras, cambio o deposito al banco -- se
        # guardaban pero no entraban en el arqueo. Resultado: cada vez que
        # recepcion sacaba plata para comprar algo, la caja "faltaba" ese
        # importe y la diferencia se atribuia a un error de conteo.
        arqueo = await db_pg.uno(
            conn,
            """select
                   coalesce((select jsonb_object_agg(method, suma)
                             from (select method, sum(amount) as suma
                                   from payments
                                   where cash_shift_id = $1 and status = 'ACTIVE'
                                   group by method) x), '{}'::jsonb) as totales,
                   coalesce((select sum(amount) from payments
                             where cash_shift_id = $1 and status = 'ACTIVE'
                               and method = 'EFECTIVO'), 0) as cobros_efectivo,
                   coalesce((select sum(case when type = 'IN' then amount else -amount end)
                             from cash_movements where cash_shift_id = $1), 0) as movimientos""",
            sid,
        )

        esperado = turno["opening_amount"] + arqueo["cobros_efectivo"] + arqueo["movimientos"]
        diferencia = round(data.counted_cash - esperado, 2)

        await conn.execute(
            """update cash_shifts set
                   status = 'CLOSED', totals = $2, counted_cash = $3,
                   difference = $4, closed_at = now(), closed_by = $5, notes = $6
               where id = $1""",
            sid, json.dumps(arqueo["totales"]), data.counted_cash,
            diferencia, id_valido(user["user_id"], "user_id"), data.notes,
        )

        if abs(diferencia) > 10:
            await create_alert(
                conn, turno["tenant_id"], "CASH_DIFFERENCE",
                AlertSeverity.WARN if abs(diferencia) < 50 else AlertSeverity.CRITICAL,
                "Diferencia en Caja",
                f"Diferencia de S/ {diferencia:.2f} al cerrar caja",
                {"cash_shift_id": shift_id}
            )

        await create_audit_log(conn, turno["tenant_id"], user["user_id"], "cash_shift",
                               "CLOSE", None, {"difference": diferencia})

    return {"message": "Caja cerrada", "difference": diferencia}

@api_router.get("/cash-shifts")
async def list_cash_shifts(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    user: dict = Depends(get_current_user)
):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        # El rango de fechas se compara contra un timestamptz real. Antes se
        # concatenaba "T23:59:59" al texto de la fecha para simular el fin del
        # dia: funcionaba de casualidad porque ISO ordena alfabeticamente, y se
        # comia los cierres ocurridos en el ultimo segundo del dia.
        return await db_pg.varias(
            conn,
            """select * from cash_shifts
               where ($1::uuid is null or tenant_id = $1)
                 and ($2::date is null or opened_at >= $2::date)
                 and ($3::date is null or opened_at < ($3::date + 1))
               order by opened_at desc""",
            tid, from_date, to_date,
        )

@api_router.post("/cash-shifts/{shift_id}/movements")
async def add_cash_movement(shift_id: str, data: CashMovementCreate, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    sid = id_valido(shift_id, "shift_id")

    async with db_pg.tx(user) as conn:
        turno = await db_pg.uno(
            conn,
            """select id, tenant_id from cash_shifts
               where id = $1 and ($2::uuid is null or tenant_id = $2) and status = 'OPEN'""",
            sid, tid,
        )
        if not turno:
            raise HTTPException(status_code=404, detail="Caja abierta no encontrada")

        nuevo_id = await conn.fetchval(
            """insert into cash_movements (tenant_id, cash_shift_id, type, amount, reason, created_by)
               values ($1, $2, $3, $4, $5, $6)
               returning id""",
            db_pg.a_uuid(turno["tenant_id"]), sid, data.type, data.amount,
            data.reason, id_valido(user["user_id"], "user_id"),
        )
    return {"id": str(nuevo_id), "message": "Movimiento registrado"}

# ============== INVOICE ENDPOINTS ==============
@api_router.post("/invoices")
async def create_invoice(data: InvoiceCreate, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    fid = id_valido(data.folio_id, "folio_id")

    if data.type == InvoiceType.FACTURA:
        if data.client_doc_type != DocType.RUC:
            raise HTTPException(status_code=400, detail="Factura requiere RUC")
        if not data.client_address:
            raise HTTPException(status_code=400, detail="Factura requiere dirección")

    es_boleta = data.type == InvoiceType.BOLETA

    # --------------------------------------------------------------------
    # Paso 1: reservar el correlativo y leer los importes.
    # --------------------------------------------------------------------
    # Va en su propia transaccion, corta, y ANTES de hablar con NubeFact. Una
    # llamada HTTP no se puede deshacer con un rollback, asi que mantener la
    # transaccion abierta mientras se espera al proveedor bloquearia la fila del
    # hotel -- y con ella toda la emision -- durante segundos.
    #
    # El correlativo se consume aunque el comprobante termine rechazado, y eso
    # es lo correcto: SUNAT exige numeracion sin saltos, y un numero emitido y
    # rechazado se resuelve con una comunicacion de baja, no reutilizandolo.
    async with db_pg.tx(user) as conn:
        cabecera = await db_pg.uno(
            conn,
            """select f.id, f.tenant_id, t.igv_rate, t.boleta_series, t.factura_series,
                      t.nubefact_ruta, t.nubefact_token, t.invoicing_mode
               from folios f
               join tenants t on t.id = f.tenant_id
               where f.id = $1 and ($2::uuid is null or f.tenant_id = $2)""",
            fid, tid,
        )
        if not cabecera:
            raise HTTPException(status_code=404, detail="Folio no encontrado")

        hotel = db_pg.a_uuid(cabecera["tenant_id"])
        series = cabecera["boleta_series"] if es_boleta else cabecera["factura_series"]
        igv_rate = cabecera["igv_rate"]

        # UPDATE ... RETURNING: incrementa y devuelve en una sola operacion
        # atomica, con la fila bloqueada. Es el equivalente exacto del
        # find_one_and_update de Mongo, pero ahora ademas el unique
        # (tenant_id, type, series, number) de la tabla invoices impide que dos
        # comprobantes acaben con el mismo numero aunque algo salga mal aqui.
        correlative = await conn.fetchval(
            f"""update tenants
                set {'boleta_correlative' if es_boleta else 'factura_correlative'}
                    = {'boleta_correlative' if es_boleta else 'factura_correlative'} + 1
                where id = $1
                returning {'boleta_correlative' if es_boleta else 'factura_correlative'}""",
            hotel,
        )

        charges = await db_pg.varias(
            conn,
            """select concept, quantity, unit_price, subtotal, igv_amount, total
               from charges where folio_id = $1 and status = 'ACTIVE'
               order by created_at""",
            fid,
        )
        importes = await db_pg.uno(
            conn,
            """select coalesce(sum(subtotal), 0)   as subtotal,
                      coalesce(sum(igv_amount), 0) as igv,
                      coalesce(sum(total), 0)      as total
               from charges where folio_id = $1 and status = 'ACTIVE'""",
            fid,
        )

    if not charges:
        raise HTTPException(status_code=400, detail="El folio no tiene cargos que facturar")

    subtotal, igv, total = importes["subtotal"], importes["igv"], importes["total"]
    tenant = {"nubefact_ruta": cabecera["nubefact_ruta"],
              "nubefact_token": cabecera["nubefact_token"],
              "invoicing_config": {"invoicing_mode": cabecera["invoicing_mode"]}}

    items = []
    for c in charges:
        items.append({
            "unidad_de_medida": "NIU",
            "codigo": "001",
            "descripcion": c.get("concept", ""),
            "cantidad": c.get("quantity", 1),
            "valor_unitario": c.get("unit_price", 0),
            # Antes iba multiplicado por 1.18 fijo, ignorando el igv_rate del
            # hotel: si alguien configuraba otra tasa, el precio unitario que
            # viajaba a SUNAT no cuadraba con el IGV declarado del propio
            # comprobante.
            "precio_unitario": round(c.get("unit_price", 0) * (1 + float(igv_rate) / 100), 2),
            "descuento": 0,
            "subtotal": c.get("subtotal", 0),
            "tipo_de_igv": 1,
            "igv": c.get("igv_amount", 0),
            "total": c.get("total", 0)
        })

    nubefact_payload = {
        "operacion": "generar_comprobante",
        "tipo_de_comprobante": 2 if data.type == InvoiceType.BOLETA else 1,
        "serie": series,
        "numero": correlative,
        "sunat_transaction": 1,
        "cliente_tipo_de_documento": {"DNI": 1, "CE": 4, "PASAPORTE": 7, "RUC": 6}.get(data.client_doc_type.value, 1),
        "cliente_numero_de_documento": data.client_doc_number,
        "cliente_denominacion": data.client_name,
        "cliente_direccion": data.client_address or "",
        "cliente_email": "",
        "fecha_de_emision": datetime.now(timezone.utc).strftime("%d-%m-%Y"),
        "moneda": 1,
        "porcentaje_de_igv": float(igv_rate),
        "total_gravada": subtotal,
        "total_igv": igv,
        "total": total,
        "items": items,
        "enviar_automaticamente_a_la_sunat": True,
        "enviar_automaticamente_al_cliente": False
    }
    
    # Paso 2: hablar con el proveedor, fuera de toda transaccion.
    nubefact_response = await NubeFactService.send_invoice(tenant, nubefact_payload)
    aceptado = bool(nubefact_response.get("success"))
    estado = InvoiceStatus.ACCEPTED.value if aceptado else InvoiceStatus.REJECTED.value

    # Paso 3: guardar el comprobante con lo que respondio SUNAT.
    async with db_pg.tx(user) as conn:
        invoice_id = await conn.fetchval(
            """insert into invoices
                   (tenant_id, folio_id, type, series, number,
                    client_doc_type, client_doc_number, client_name, client_address,
                    subtotal, igv, total, status,
                    nubefact_request, nubefact_response,
                    pdf_url, xml_url, cdr_url, hash, qr, issued_by)
               values ($1, $2, $3::invoice_type, $4, $5,
                       $6::doc_type, $7, $8, $9,
                       $10, $11, $12, $13::invoice_status,
                       $14, $15, $16, $17, $18, $19, $20, $21)
               returning id""",
            hotel, fid, data.type.value, series, correlative,
            data.client_doc_type.value, data.client_doc_number,
            data.client_name, data.client_address,
            subtotal, igv, total, estado,
            json.dumps(nubefact_payload), json.dumps(nubefact_response),
            nubefact_response.get("enlace_del_pdf"),
            nubefact_response.get("enlace_del_xml"),
            nubefact_response.get("enlace_del_cdr"),
            nubefact_response.get("hash"), nubefact_response.get("qr"),
            id_valido(user["user_id"], "user_id"),
        )

        await create_audit_log(conn, hotel, user["user_id"], "invoice", "CREATE", None,
                               {"series": series, "number": correlative, "type": data.type.value})

        if not aceptado:
            await create_alert(
                conn, hotel, "INVOICE_REJECTED", AlertSeverity.CRITICAL,
                "Comprobante Rechazado",
                f"{data.type.value} {series}-{correlative} fue rechazada: "
                f"{nubefact_response.get('error', 'Error desconocido')}",
                {"invoice_id": str(invoice_id)}
            )

    return {
        "id": str(invoice_id),
        "series": series,
        "number": correlative,
        "status": estado,
        "message": "Comprobante emitido" if aceptado else "Error al emitir comprobante"
    }

@api_router.get("/invoices")
async def list_invoices(
    type: Optional[InvoiceType] = None,
    status: Optional[InvoiceStatus] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    search: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        return await db_pg.varias(
            conn,
            """select * from invoices
               where ($1::uuid is null or tenant_id = $1)
                 and ($2::text is null or type::text = $2)
                 and ($3::text is null or status::text = $3)
                 and ($4::date is null or issued_at >= $4::date)
                 and ($5::date is null or issued_at < ($5::date + 1))
                 and ($6::text is null
                      or client_name       ilike '%' || $6 || '%'
                      or client_doc_number ilike '%' || $6 || '%')
               order by issued_at desc""",
            tid,
            type.value if type else None,
            status.value if status else None,
            from_date, to_date, search,
        )

@api_router.get("/invoices/{invoice_id}")
async def get_invoice(invoice_id: str, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        invoice = await db_pg.uno(
            conn,
            "select * from invoices where id = $1 and ($2::uuid is null or tenant_id = $2)",
            id_valido(invoice_id, "invoice_id"), tid,
        )
    if not invoice:
        raise HTTPException(status_code=404, detail="Comprobante no encontrado")
    return invoice

@api_router.post("/invoices/{invoice_id}/void")
async def void_invoice(invoice_id: str, data: VoidRequest, user: dict = Depends(require_roles(Role.ADMIN))):
    tid = tenant_de(user)
    iid = id_valido(invoice_id, "invoice_id")

    async with db_pg.tx(user) as conn:
        # `and status <> 'VOIDED'` dentro del WHERE: dos anulaciones simultaneas
        # ya no pisan la fecha ni el motivo de la primera.
        anulado = await db_pg.uno(
            conn,
            """update invoices
               set status = 'VOIDED', void_reason = $3, voided_at = now()
               where id = $1 and ($2::uuid is null or tenant_id = $2)
                 and status <> 'VOIDED'
               returning tenant_id""",
            iid, tid, data.reason,
        )
        if not anulado:
            existe = await conn.fetchval(
                "select status from invoices where id = $1 and ($2::uuid is null or tenant_id = $2)",
                iid, tid,
            )
            if existe:
                raise HTTPException(status_code=400, detail="Comprobante ya anulado")
            raise HTTPException(status_code=404, detail="Comprobante no encontrado")

        await create_audit_log(conn, anulado["tenant_id"], user["user_id"], "invoice",
                               "VOID", {"id": invoice_id}, {"reason": data.reason})

    return {"message": "Comprobante anulado"}

# ============== HOUSEKEEPING ENDPOINTS ==============
@api_router.get("/housekeeping/tasks")
async def list_housekeeping_tasks(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        return await db_pg.varias(
            conn,
            """select t.*,
                      jsonb_build_object('number', h.number, 'floor', h.floor) as room
               from housekeeping_tasks t
               join rooms h on h.id = t.room_id
               where ($1::uuid is null or t.tenant_id = $1)
                 and ($2::text is null or t.status = $2)
                 and ($3::text is null or t.priority = $3)
               order by t.created_at desc""",
            tid, status, priority,
        )

@api_router.post("/housekeeping/tasks/{task_id}/complete")
async def complete_housekeeping_task(task_id: str, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    tkid = id_valido(task_id, "task_id")
    quien = id_valido(user["user_id"], "user_id")

    async with db_pg.tx(user) as conn:
        # `and status <> 'DONE'` evita que completar dos veces la misma tarea
        # deje dos entradas en la bitacora de limpieza.
        tarea = await db_pg.uno(
            conn,
            """update housekeeping_tasks
               set status = 'DONE', completed_by = $3, completed_at = now()
               where id = $1 and ($2::uuid is null or tenant_id = $2)
                 and status <> 'DONE'
               returning tenant_id, room_id""",
            tkid, tid, quien,
        )
        if not tarea:
            existe = await conn.fetchval(
                "select 1 from housekeeping_tasks where id = $1 and ($2::uuid is null or tenant_id = $2)",
                tkid, tid,
            )
            if existe:
                return {"message": "Tarea completada"}
            raise HTTPException(status_code=404, detail="Tarea no encontrada")

        # El estado anterior de la habitacion se lee del propio UPDATE en vez de
        # asumir que era DIRTY: una habitacion puede venir de INSPECT o de
        # CLEANING, y la bitacora registraba una transicion falsa.
        cambio = await db_pg.uno(
            conn,
            """update rooms r
               set housekeeping_status = 'CLEAN'
               from rooms viejo
               where r.id = viejo.id and r.id = $1
               returning viejo.housekeeping_status as antes""",
            db_pg.a_uuid(tarea["room_id"]),
        )
        await conn.execute(
            """insert into housekeeping_logs
                   (tenant_id, room_id, from_housekeeping, to_housekeeping, by_user)
               values ($1, $2, $3::housekeeping_state, 'CLEAN', $4)""",
            db_pg.a_uuid(tarea["tenant_id"]), db_pg.a_uuid(tarea["room_id"]),
            cambio["antes"], quien,
        )

    return {"message": "Tarea completada"}

@api_router.get("/housekeeping/board")
async def get_housekeeping_board(user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        # El agrupado por piso lo hace Postgres. La respuesta mantiene la forma
        # {"floors": {piso: [habitaciones]}} que espera el tablero.
        filas = await db_pg.varias(
            conn,
            """select floor::text as piso,
                      jsonb_agg(to_jsonb(r.*) order by r.number) as habitaciones
               from rooms r
               where ($1::uuid is null or tenant_id = $1) and is_active
               group by floor
               order by floor""",
            tid,
        )
    return {"floors": {f["piso"]: f["habitaciones"] for f in filas}}

# ============== MAINTENANCE ENDPOINTS ==============
@api_router.post("/maintenance/tickets")
async def create_maintenance_ticket(data: MaintenanceTicketCreate, user: dict = Depends(get_current_user)):
    tid = db_pg.a_uuid(user["tenant_id"], "tenant_id")
    hid = id_valido(data.room_id, "room_id")

    async with db_pg.tx(user) as conn:
        try:
            nuevo_id = await conn.fetchval(
                """insert into maintenance_tickets
                       (tenant_id, room_id, title, description, priority, estimated_cost, created_by)
                   values ($1, $2, $3, $4, $5, $6, $7)
                   returning id""",
                tid, hid, data.title, data.description, data.priority,
                data.estimated_cost, id_valido(user["user_id"], "user_id"),
            )
        except db_pg.ForeignKeyViolationError:
            raise HTTPException(status_code=404, detail="Habitación no encontrada")

        # Un ticket critico saca la habitacion de servicio. Va en la misma
        # transaccion que el alta del ticket: antes eran tres escrituras
        # sueltas, y si fallaba la segunda quedaba un ticket critico con la
        # habitacion todavia disponible para vender.
        if data.priority == "CRITICAL":
            await conn.execute(
                "update rooms set housekeeping_status = 'OUT_OF_ORDER' where id = $1 and tenant_id = $2",
                hid, tid,
            )
            await create_alert(
                conn, tid, "ROOM_OUT_OF_ORDER", AlertSeverity.CRITICAL,
                "Habitación Fuera de Servicio",
                "Habitación marcada como fuera de servicio por ticket de mantenimiento crítico",
                {"ticket_id": str(nuevo_id), "room_id": data.room_id}
            )

    return {"id": str(nuevo_id), "message": "Ticket creado"}

@api_router.get("/maintenance/tickets")
async def list_maintenance_tickets(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        return await db_pg.varias(
            conn,
            """select t.*,
                      jsonb_build_object('number', h.number, 'floor', h.floor) as room
               from maintenance_tickets t
               join rooms h on h.id = t.room_id
               where ($1::uuid is null or t.tenant_id = $1)
                 and ($2::text is null or t.status = $2)
                 and ($3::text is null or t.priority = $3)
               order by t.created_at desc""",
            tid, status, priority,
        )

@api_router.put("/maintenance/tickets/{ticket_id}")
async def update_maintenance_ticket(
    ticket_id: str,
    status: Optional[str] = None,
    assigned_to: Optional[str] = None,
    actual_cost: Optional[float] = None,
    user: dict = Depends(get_current_user)
):
    tid = tenant_de(user)
    tkid = id_valido(ticket_id, "ticket_id")

    async with db_pg.tx(user) as conn:
        ticket = await db_pg.uno(
            conn,
            """update maintenance_tickets set
                   status      = coalesce($3::text, status),
                   assigned_to = coalesce($4::uuid, assigned_to),
                   actual_cost = coalesce($5::numeric, actual_cost),
                   resolved_at = case when $3 = 'RESOLVED' then now() else resolved_at end
               where id = $1 and ($2::uuid is null or tenant_id = $2)
               returning priority, room_id, tenant_id""",
            tkid, tid, status, id_valido(assigned_to, "assigned_to"), actual_cost,
        )
        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket no encontrado")

        # Al resolver un ticket critico la habitacion vuelve del limbo, pero
        # queda SUCIA y no limpia: nadie la ha aseado todavia, y darla por
        # limpia la haria asignable a un huesped que la encontraria como estaba.
        if status == "RESOLVED" and ticket["priority"] == "CRITICAL" and ticket.get("room_id"):
            await conn.execute(
                """update rooms set housekeeping_status = 'DIRTY'
                   where id = $1 and housekeeping_status = 'OUT_OF_ORDER'""",
                db_pg.a_uuid(ticket["room_id"]),
            )

    return {"message": "Ticket actualizado"}

# ============== ALERT ENDPOINTS ==============
@api_router.get("/alerts")
async def list_alerts(
    status: Optional[str] = None,
    severity: Optional[AlertSeverity] = None,
    user: dict = Depends(get_current_user)
):
    tid = tenant_de(user)
    # El personal de limpieza solo ve las alertas de su area.
    tipos_limpieza = ["DIRTY_ROOM", "ROOM_OUT_OF_ORDER"] \
        if user["role"] == Role.HOUSEKEEPING.value else None

    async with db_pg.tx(user) as conn:
        return await db_pg.varias(
            conn,
            """select * from alerts
               where ($1::uuid is null or tenant_id = $1)
                 and ($2::text is null or status = $2)
                 and ($3::text is null or severity::text = $3)
                 and ($4::text[] is null or type = any($4))
               order by created_at desc""",
            tid, status, severity.value if severity else None, tipos_limpieza,
        )

@api_router.post("/alerts/{alert_id}/resolve")
async def resolve_alert(alert_id: str, data: AlertResolve, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        # Antes el update no comprobaba si habia acertado: resolver una alerta
        # inexistente o de otro hotel devolvia "Alerta resuelta" igual.
        resuelta = await conn.fetchval(
            """update alerts
               set status = 'RESOLVED', resolved_by = $3, resolved_at = now(), notes = $4
               where id = $1 and ($2::uuid is null or tenant_id = $2)
               returning id""",
            id_valido(alert_id, "alert_id"), tid,
            id_valido(user["user_id"], "user_id"), data.notes,
        )
    if not resuelta:
        raise HTTPException(status_code=404, detail="Alerta no encontrada")
    return {"message": "Alerta resuelta"}

# ============== PRODUCTS/SERVICES ENDPOINTS ==============
@api_router.post("/products")
async def create_product(name: str = Body(...), category: str = Body(...), unit_price: float = Body(...), tax_type: str = Body("IGV"), user: dict = Depends(require_roles(Role.ADMIN))):
    async with db_pg.tx(user) as conn:
        nuevo_id = await conn.fetchval(
            """insert into products (tenant_id, name, category, unit_price, tax_type)
               values ($1, $2, $3, $4, $5)
               returning id""",
            db_pg.a_uuid(user["tenant_id"]), name, category, unit_price, tax_type,
        )
    return {"id": str(nuevo_id), "message": "Producto creado"}

@api_router.get("/products")
async def list_products(category: Optional[str] = None, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        return await db_pg.varias(
            conn,
            """select * from products
               where ($1::uuid is null or tenant_id = $1) and is_active
                 and ($2::text is null or category = $2)
               order by name""",
            tid, category,
        )

# ============== DASHBOARD ENDPOINTS ==============
@api_router.get("/dashboard/kpis")
async def get_dashboard_kpis(user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    today = datetime.now(timezone.utc).date()
    month_start = today.replace(day=1)

    async with db_pg.tx(user) as conn:
        # Los ocho count_documents y los cuatro pipelines de agregacion se
        # resuelven en UNA consulta: cada subselect es independiente y Postgres
        # los ejecuta en el mismo viaje. El dashboard es lo primero que se abre
        # cada manana y antes disparaba doce consultas seguidas.
        #
        # Las fechas se comparan como fechas, no como prefijos de texto: antes
        # "created_at >= '2026-08-01'" funcionaba solo porque ISO ordena
        # alfabeticamente igual que cronologicamente.
        k = await db_pg.uno(
            conn,
            """select
                 (select count(*) from rooms
                   where ($1::uuid is null or tenant_id = $1) and is_active) as rooms_total,
                 (select count(*) from rooms
                   where ($1::uuid is null or tenant_id = $1) and occupancy_status = 'OCCUPIED') as rooms_occupied,
                 (select count(*) from rooms
                   where ($1::uuid is null or tenant_id = $1) and housekeeping_status = 'DIRTY') as rooms_dirty,
                 (select count(*) from rooms
                   where ($1::uuid is null or tenant_id = $1) and housekeeping_status = 'OUT_OF_ORDER') as rooms_ooo,
                 (select count(*) from reservations
                   where ($1::uuid is null or tenant_id = $1)
                     and checkin_date = $2 and status = 'CONFIRMED') as arrivals,
                 (select count(*) from reservations
                   where ($1::uuid is null or tenant_id = $1)
                     and checkout_date = $2 and status = 'CHECKED_IN') as departures,
                 (select coalesce(sum(amount), 0) from payments
                   where ($1::uuid is null or tenant_id = $1)
                     and status = 'ACTIVE' and created_at >= $2::date) as revenue_today,
                 (select coalesce(sum(balance), 0) from folios
                   where ($1::uuid is null or tenant_id = $1)
                     and status = 'OPEN' and balance > 0) as outstanding,
                 (select coalesce(sum(amount), 0) from payments
                   where ($1::uuid is null or tenant_id = $1)
                     and status = 'ACTIVE' and created_at >= $3::date) as revenue_month,
                 (select count(*) from reservations
                   where ($1::uuid is null or tenant_id = $1)
                     and created_at >= $3::date and status = 'CANCELLED') as cancellations,
                 (select count(*) from reservations
                   where ($1::uuid is null or tenant_id = $1)
                     and created_at >= $3::date and status = 'NO_SHOW') as no_shows,
                 (select coalesce(sum(total), 0) from charges
                   where ($1::uuid is null or tenant_id = $1)
                     and created_at >= $3::date and category = 'HABITACION'
                     and status = 'ACTIVE') as ingresos_habitacion,
                 (select coalesce(sum(quantity), 0) from charges
                   where ($1::uuid is null or tenant_id = $1)
                     and created_at >= $3::date and category = 'HABITACION'
                     and status = 'ACTIVE') as noches_vendidas""",
            tid, today, month_start,
        )

    rooms_total = k["rooms_total"]
    rooms_occupied = k["rooms_occupied"]
    rooms_dirty = k["rooms_dirty"]
    rooms_ooo = k["rooms_ooo"]
    arrivals_today = k["arrivals"]
    departures_today = k["departures"]
    revenue_today = k["revenue_today"]
    outstanding = k["outstanding"]
    revenue_month = k["revenue_month"]
    cancellations = k["cancellations"]
    no_shows = k["no_shows"]

    # ADR = ingreso medio por noche de habitacion vendida.
    room_nights_sold = k["noches_vendidas"]
    adr = (k["ingresos_habitacion"] / room_nights_sold) if room_nights_sold else 0

    occupancy_rate = (rooms_occupied / rooms_total * 100) if rooms_total > 0 else 0
    revpar = (revenue_month / (rooms_total * today.day)) if rooms_total > 0 else 0

    return {
        "today": {
            "occupancy_rate": round(occupancy_rate, 1),
            "rooms_occupied": rooms_occupied,
            "rooms_total": rooms_total,
            "rooms_dirty": rooms_dirty,
            "rooms_ooo": rooms_ooo,
            "arrivals": arrivals_today,
            "departures": departures_today,
            "revenue": round(revenue_today, 2),
            "outstanding": round(outstanding, 2)
        },
        "month": {
            "revenue": round(revenue_month, 2),
            "adr": round(adr, 2),
            "revpar": round(revpar, 2),
            "cancellations": cancellations,
            "no_shows": no_shows,
            "room_nights_sold": room_nights_sold
        }
    }

@api_router.get("/dashboard/charts/revenue")
async def get_revenue_chart(days: int = 30, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        # El agrupado por dia sale de date_trunc sobre un timestamptz real, en
        # vez de recortar los 10 primeros caracteres del texto de la fecha.
        return await db_pg.varias(
            conn,
            """select to_char(created_at::date, 'YYYY-MM-DD') as date,
                      sum(total) as total,
                      sum(total) filter (where category  = 'HABITACION') as rooms,
                      sum(total) filter (where category <> 'HABITACION') as extras
               from charges
               where ($1::uuid is null or tenant_id = $1)
                 and status = 'ACTIVE'
                 and created_at >= now() - ($2 || ' days')::interval
               group by created_at::date
               order by created_at::date""",
            tid, str(days),
        )

@api_router.get("/dashboard/charts/occupancy")
async def get_occupancy_chart(days: int = 30, user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        # generate_series produce los dias del periodo y un LATERAL cuenta las
        # habitaciones ocupadas en cada uno. Antes era un bucle en Python con
        # una consulta por dia: 31 viajes a la base para pintar un grafico de
        # un mes, y 91 si alguien pedia el trimestre.
        return await db_pg.varias(
            conn,
            """with total as (
                   select count(*)::numeric as n from rooms
                   where ($1::uuid is null or tenant_id = $1) and is_active
               )
               select to_char(d::date, 'YYYY-MM-DD') as date,
                      o.ocupadas as occupied,
                      case when t.n > 0
                           then round(o.ocupadas / t.n * 100, 1)
                           else 0 end as rate
               from generate_series(
                        current_date - ($2::int - 1),
                        current_date,
                        interval '1 day') as d
               cross join total t
               left join lateral (
                   select count(*)::numeric as ocupadas
                   from reservations r
                   where ($1::uuid is null or r.tenant_id = $1)
                     and r.status in ('CHECKED_IN', 'CHECKED_OUT')
                     and r.checkin_date  <= d::date
                     and r.checkout_date >  d::date
               ) o on true
               order by d""",
            tid, days,
        )

@api_router.get("/dashboard/charts/payment-methods")
async def get_payment_methods_chart(user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        return await db_pg.varias(
            conn,
            """select method::text as method, sum(amount) as total
               from payments
               where ($1::uuid is null or tenant_id = $1)
                 and status = 'ACTIVE'
                 and created_at >= date_trunc('month', current_date)
               group by method
               order by total desc""",
            tid,
        )

@api_router.get("/dashboard/charts/room-status")
async def get_room_status_chart(user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        # La clasificacion en las cuatro categorias del grafico la hace la
        # consulta, en lugar de traer los grupos crudos y re-agregarlos con
        # ifs en Python. El orden de los CASE importa: fuera de servicio manda
        # sobre todo lo demas, y ocupada sobre el estado de limpieza.
        return await db_pg.uno(
            conn,
            """select
                   count(*) filter (where housekeeping_status = 'OUT_OF_ORDER') as out_of_order,
                   count(*) filter (where housekeeping_status <> 'OUT_OF_ORDER'
                                      and occupancy_status = 'OCCUPIED')        as occupied,
                   count(*) filter (where housekeeping_status = 'CLEAN'
                                      and occupancy_status <> 'OCCUPIED')       as vacant_clean,
                   count(*) filter (where housekeeping_status not in ('OUT_OF_ORDER', 'CLEAN')
                                      and occupancy_status <> 'OCCUPIED')       as vacant_dirty
               from rooms
               where ($1::uuid is null or tenant_id = $1) and is_active""",
            tid,
        )

@api_router.get("/dashboard/charts/invoicing-status")
async def get_invoicing_status_chart(user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        return await db_pg.varias(
            conn,
            """select status::text as status, count(*) as count, sum(total) as total
               from invoices
               where ($1::uuid is null or tenant_id = $1)
                 and issued_at >= date_trunc('month', current_date)
               group by status""",
            tid,
        )

@api_router.get("/dashboard/charts/top-products")
async def get_top_products_chart(user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        return await db_pg.varias(
            conn,
            """select concept as product, sum(total) as total, sum(quantity) as quantity
               from charges
               where ($1::uuid is null or tenant_id = $1)
                 and status = 'ACTIVE'
                 and category <> 'HABITACION'
                 and created_at >= date_trunc('month', current_date)
               group by concept
               order by total desc
               limit 10""",
            tid,
        )

# ============== REPORTS ENDPOINTS ==============
@api_router.get("/reports/monthly-occupancy")
async def get_monthly_occupancy_report(month: int = None, year: int = None, user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).date()
    
    if not month:
        month = today.month
    if not year:
        year = today.year
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    
    tid = tenant_de(user)
    days_in_month = (end_date - start_date).days

    async with db_pg.tx(user) as conn:
        r = await db_pg.uno(
            conn,
            """select
                 (select count(*) from rooms
                   where ($1::uuid is null or tenant_id = $1) and is_active) as rooms_total,

                 -- Noches-habitacion vendidas DE VERDAD. Antes era el numero de
                 -- reservas del mes: una estadia de cinco noches contaba como
                 -- una, asi que la ocupacion salia hasta cinco veces mas baja
                 -- de lo real y el ADR y el RevPAR heredaban el error.
                 -- Aca se suman los dias que cada reserva pasa DENTRO del mes,
                 -- recortando la parte que se sale por cualquiera de los dos
                 -- extremos.
                 (select coalesce(sum(
                            least(checkout_date, $3::date) - greatest(checkin_date, $2::date)
                         ), 0)
                    from reservations
                   where ($1::uuid is null or tenant_id = $1)
                     and status in ('CHECKED_IN', 'CHECKED_OUT')
                     and checkin_date  <  $3::date
                     and checkout_date >  $2::date) as room_nights_sold,

                 (select count(*) from reservations
                   where ($1::uuid is null or tenant_id = $1)
                     and checkin_date >= $2::date and checkin_date < $3::date
                     and status in ('CHECKED_IN', 'CHECKED_OUT')) as checkins,
                 (select count(*) from reservations
                   where ($1::uuid is null or tenant_id = $1)
                     and checkout_date >= $2::date and checkout_date < $3::date
                     and status = 'CHECKED_OUT') as checkouts,
                 (select count(*) from reservations
                   where ($1::uuid is null or tenant_id = $1)
                     and created_at >= $2::date and created_at < $3::date
                     and status = 'CANCELLED') as cancellations,
                 (select count(*) from reservations
                   where ($1::uuid is null or tenant_id = $1)
                     and checkin_date >= $2::date and checkin_date < $3::date
                     and status = 'NO_SHOW') as no_shows,
                 (select coalesce(sum(total), 0) from charges
                   where ($1::uuid is null or tenant_id = $1)
                     and created_at >= $2::date and created_at < $3::date
                     and category = 'HABITACION' and status = 'ACTIVE') as room_revenue,
                 (select coalesce(sum(quantity), 0) from charges
                   where ($1::uuid is null or tenant_id = $1)
                     and created_at >= $2::date and created_at < $3::date
                     and category = 'HABITACION' and status = 'ACTIVE') as actual_nights""",
            tid, start_date, end_date,
        )

    rooms_total = r["rooms_total"]
    room_nights_available = rooms_total * days_in_month
    room_nights_sold = r["room_nights_sold"]
    checkins = r["checkins"]
    checkouts = r["checkouts"]
    cancellations = r["cancellations"]
    no_shows = r["no_shows"]
    room_revenue = r["room_revenue"]
    actual_nights = r["actual_nights"]

    occupancy_avg = (room_nights_sold / room_nights_available * 100) if room_nights_available > 0 else 0
    adr = (room_revenue / actual_nights) if actual_nights > 0 else 0
    revpar = (room_revenue / room_nights_available) if room_nights_available > 0 else 0
    
    return {
        "period": {"month": month, "year": year},
        "summary": {
            "occupancy_avg": round(occupancy_avg, 1),
            "room_nights_available": room_nights_available,
            "room_nights_sold": room_nights_sold,
            "checkins": checkins,
            "checkouts": checkouts,
            "cancellations": cancellations,
            "no_shows": no_shows,
            "adr": round(adr, 2),
            "revpar": round(revpar, 2),
            "room_revenue": round(room_revenue, 2)
        }
    }

@api_router.get("/reports/monthly-revenue")
async def get_monthly_revenue_report(month: int = None, year: int = None, user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).date()
    
    if not month:
        month = today.month
    if not year:
        year = today.year
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        cat = await db_pg.varias(
            conn,
            """select category, sum(total) as total
               from charges
               where ($1::uuid is null or tenant_id = $1)
                 and status = 'ACTIVE'
                 and created_at >= $2::date and created_at < $3::date
               group by category""",
            tid, start_date, end_date,
        )
        met = await db_pg.varias(
            conn,
            """select method::text as method, sum(amount) as total
               from payments
               where ($1::uuid is null or tenant_id = $1)
                 and status = 'ACTIVE'
                 and created_at >= $2::date and created_at < $3::date
               group by method""",
            tid, start_date, end_date,
        )

    by_category = {c["category"]: c["total"] for c in cat}
    by_method = {m["method"]: m["total"] for m in met}

    total_charges = sum(by_category.values())
    total_payments = sum(by_method.values())
    
    return {
        "period": {"month": month, "year": year},
        "summary": {
            "total_charges": round(total_charges, 2),
            "total_payments": round(total_payments, 2),
            "rooms_revenue": round(by_category.get("HABITACION", 0), 2),
            "extras_revenue": round(total_charges - by_category.get("HABITACION", 0), 2)
        },
        "by_category": by_category,
        "by_payment_method": by_method
    }

@api_router.get("/reports/monthly-invoicing")
async def get_monthly_invoicing_report(month: int = None, year: int = None, user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).date()
    
    if not month:
        month = today.month
    if not year:
        year = today.year
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        # Los dos agrupados recorren exactamente las mismas filas, asi que se
        # leen de una sola pasada en vez de dos.
        filas = await db_pg.varias(
            conn,
            """select type::text as clave, 'type' as eje, count(*) as count, sum(total) as total
               from invoices
               where ($1::uuid is null or tenant_id = $1)
                 and issued_at >= $2::date and issued_at < $3::date
               group by type
               union all
               select status::text, 'status', count(*), sum(total)
               from invoices
               where ($1::uuid is null or tenant_id = $1)
                 and issued_at >= $2::date and issued_at < $3::date
               group by status""",
            tid, start_date, end_date,
        )

    by_type = {f["clave"]: {"count": f["count"], "total": f["total"]}
               for f in filas if f["eje"] == "type"}
    by_status = {f["clave"]: {"count": f["count"], "total": f["total"]}
                 for f in filas if f["eje"] == "status"}

    total_invoices = sum(d.get("count", 0) for d in by_type.values())
    total_amount = sum(d.get("total", 0) for d in by_type.values())
    
    return {
        "period": {"month": month, "year": year},
        "summary": {
            "total_invoices": total_invoices,
            "total_amount": round(total_amount, 2)
        },
        "by_type": by_type,
        "by_status": by_status
    }

# ============== EXPORT ENDPOINTS ==============
@api_router.get("/reports/export/excel")
async def export_report_excel(
    report_type: str = Query(..., enum=["occupancy", "revenue", "invoicing"]),
    month: int = None,
    year: int = None,
    user: dict = Depends(get_current_user)
):
    """Export report to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    today = datetime.now(timezone.utc).date()
    if not month:
        month = today.month
    if not year:
        year = today.year
    
    # Get report data
    if report_type == "occupancy":
        report_data = await get_monthly_occupancy_report(month, year, user)
        filename = f"ocupacion_{year}_{month:02d}.xlsx"
    elif report_type == "revenue":
        report_data = await get_monthly_revenue_report(month, year, user)
        filename = f"ingresos_{year}_{month:02d}.xlsx"
    else:
        report_data = await get_monthly_invoicing_report(month, year, user)
        filename = f"facturacion_{year}_{month:02d}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    months_es = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Reporte de {report_type.title()} - {months_es[month]} {year}"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A3'] = "Período:"
    ws['B3'] = f"{months_es[month]} {year}"
    
    row = 5
    
    # Summary section
    ws[f'A{row}'] = "Resumen"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    summary = report_data.get("summary", {})
    for key, value in summary.items():
        ws[f'A{row}'] = key.replace("_", " ").title()
        ws[f'B{row}'] = f"S/ {value:.2f}" if isinstance(value, float) else value
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    row += 2
    
    # Category breakdown if available
    if "by_category" in report_data:
        ws[f'A{row}'] = "Por Categoría"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = "Categoría"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'] = "Total"
        ws[f'B{row}'].font = header_font
        ws[f'B{row}'].fill = header_fill
        row += 1
        for key, value in report_data["by_category"].items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = f"S/ {value:.2f}"
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
    
    # Payment method breakdown if available
    if "by_payment_method" in report_data:
        row += 1
        ws[f'A{row}'] = "Por Método de Pago"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = "Método"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'] = "Total"
        ws[f'B{row}'].font = header_font
        ws[f'B{row}'].fill = header_fill
        row += 1
        for key, value in report_data["by_payment_method"].items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = f"S/ {value:.2f}"
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
    
    # Invoice type breakdown if available
    if "by_type" in report_data:
        row += 1
        ws[f'A{row}'] = "Por Tipo de Comprobante"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = "Tipo"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'] = "Cantidad"
        ws[f'B{row}'].font = header_font
        ws[f'B{row}'].fill = header_fill
        ws[f'C{row}'] = "Total"
        ws[f'C{row}'].font = header_font
        ws[f'C{row}'].fill = header_fill
        row += 1
        for key, data in report_data["by_type"].items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = data.get("count", 0)
            ws[f'C{row}'] = f"S/ {data.get('total', 0):.2f}"
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            ws[f'C{row}'].border = border
            row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/export/pdf")
async def export_report_pdf(
    report_type: str = Query(..., enum=["occupancy", "revenue", "invoicing"]),
    month: int = None,
    year: int = None,
    user: dict = Depends(get_current_user)
):
    """Export report to PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    today = datetime.now(timezone.utc).date()
    if not month:
        month = today.month
    if not year:
        year = today.year
    
    # Get report data
    if report_type == "occupancy":
        report_data = await get_monthly_occupancy_report(month, year, user)
        title = "Reporte de Ocupación"
        filename = f"ocupacion_{year}_{month:02d}.pdf"
    elif report_type == "revenue":
        report_data = await get_monthly_revenue_report(month, year, user)
        title = "Reporte de Ingresos"
        filename = f"ingresos_{year}_{month:02d}.pdf"
    else:
        report_data = await get_monthly_invoicing_report(month, year, user)
        title = "Reporte de Facturación"
        filename = f"facturacion_{year}_{month:02d}.pdf"
    
    months_es = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=20)
    subtitle_style = ParagraphStyle('SubTitle', parent=styles['Heading2'], fontSize=12, spaceAfter=10)
    
    elements = []
    
    # Title
    elements.append(Paragraph(f"{title} - {months_es[month]} {year}", title_style))
    elements.append(Spacer(1, 20))
    
    # Summary table
    elements.append(Paragraph("Resumen", subtitle_style))
    summary = report_data.get("summary", {})
    summary_data = [["Métrica", "Valor"]]
    for key, value in summary.items():
        display_value = f"S/ {value:.2f}" if isinstance(value, float) else str(value)
        summary_data.append([key.replace("_", " ").title(), display_value])
    
    summary_table = Table(summary_data, colWidths=[200, 150])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWHEIGHTS', (0, 0), (-1, -1), 25),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # By category if available
    if "by_category" in report_data and report_data["by_category"]:
        elements.append(Paragraph("Por Categoría", subtitle_style))
        cat_data = [["Categoría", "Total"]]
        for key, value in report_data["by_category"].items():
            cat_data.append([key, f"S/ {value:.2f}"])
        cat_table = Table(cat_data, colWidths=[200, 150])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ROWHEIGHTS', (0, 0), (-1, -1), 22),
        ]))
        elements.append(cat_table)
        elements.append(Spacer(1, 20))
    
    # By payment method if available
    if "by_payment_method" in report_data and report_data["by_payment_method"]:
        elements.append(Paragraph("Por Método de Pago", subtitle_style))
        pm_data = [["Método", "Total"]]
        for key, value in report_data["by_payment_method"].items():
            pm_data.append([key, f"S/ {value:.2f}"])
        pm_table = Table(pm_data, colWidths=[200, 150])
        pm_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ROWHEIGHTS', (0, 0), (-1, -1), 22),
        ]))
        elements.append(pm_table)
        elements.append(Spacer(1, 20))
    
    # By type if available (invoicing)
    if "by_type" in report_data and report_data["by_type"]:
        elements.append(Paragraph("Por Tipo de Comprobante", subtitle_style))
        type_data = [["Tipo", "Cantidad", "Total"]]
        for key, data in report_data["by_type"].items():
            type_data.append([key, str(data.get("count", 0)), f"S/ {data.get('total', 0):.2f}"])
        type_table = Table(type_data, colWidths=[150, 100, 100])
        type_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ROWHEIGHTS', (0, 0), (-1, -1), 22),
        ]))
        elements.append(type_table)
    
    # Footer
    elements.append(Spacer(1, 40))
    elements.append(Paragraph(f"Generado el: {today.strftime('%d/%m/%Y')}", styles['Normal']))
    
    doc.build(elements)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============== WALK-IN ENDPOINT ==============
@api_router.post("/reservations/walkin")
async def create_walkin(
    guest_data: GuestCreate,
    room_id: str = Body(...),
    checkout_date: date = Body(...),
    adults: int = Body(1),
    children: int = Body(0),
    notes: str = Body(None),
    user: dict = Depends(get_current_user)
):
    """Registra un huesped sin reserva previa y lo ingresa en el acto."""
    tid = db_pg.a_uuid(user["tenant_id"], "tenant_id")
    hid = id_valido(room_id, "room_id")
    quien = id_valido(user["user_id"], "user_id")

    today = datetime.now(timezone.utc).date()
    nights = (checkout_date - today).days
    if nights < 1:
        raise HTTPException(status_code=400, detail="Fecha de checkout debe ser posterior a hoy")

    # Todo el walk-in -- huesped, reserva, estadia, folio y estado de la
    # habitacion -- en UNA transaccion. Antes eran siete escrituras sueltas con
    # el huesped delante del mostrador: si fallaba a mitad quedaba una
    # habitacion marcada como ocupada sin folio donde cargarle nada.
    async with db_pg.tx(user) as conn:
        await conn.execute("select pg_advisory_xact_lock(hashtext($1))", str(tid))

        habitacion = await db_pg.uno(
            conn,
            """select h.id, h.number, h.room_type_id, h.occupancy_status,
                      h.housekeeping_status, rt.base_price
               from rooms h
               join room_types rt on rt.id = h.room_type_id
               where h.id = $1 and h.tenant_id = $2""",
            hid, tid,
        )
        if not habitacion:
            raise HTTPException(status_code=404, detail="Habitación no encontrada")
        if habitacion["occupancy_status"] != OccupancyStatus.VACANT.value:
            raise HTTPException(status_code=400, detail="Habitación no disponible")
        if habitacion["housekeeping_status"] == HousekeepingStatus.OUT_OF_ORDER.value:
            raise HTTPException(status_code=400, detail="Habitación fuera de servicio")

        total_estimated = nights * habitacion["base_price"]

        # Mismo ON CONFLICT que en el alta normal de huespedes: si el DNI ya
        # existe se reutiliza la ficha en vez de duplicarla.
        guest_id = await conn.fetchval(
            """insert into guests (tenant_id, doc_type, doc_number, full_name,
                                   phone, email, nationality, address)
               values ($1, $2::doc_type, $3, $4, $5, $6, $7, $8)
               on conflict (tenant_id, doc_type, doc_number) do update
                   set full_name = excluded.full_name
               returning id""",
            tid, guest_data.doc_type.value, guest_data.doc_number, guest_data.full_name,
            guest_data.phone, guest_data.email, guest_data.nationality, guest_data.address,
        )

        # El correlativo sale del maximo real de los codigos WLK-, no de contar
        # todas las reservas del hotel: antes un walk-in podia recibir un numero
        # ya usado por otro walk-in si entremedio se cancelaba una reserva.
        n = await conn.fetchval(
            """select coalesce(max(substring(code from 'WLK-([0-9]+)$')::int), 0) + 1
               from reservations where tenant_id = $1""",
            tid,
        )
        code = f"WLK-{n:06d}"

        try:
            reservation_id = await conn.fetchval(
                """insert into reservations
                       (tenant_id, code, guest_id, room_type_id, room_id,
                        checkin_date, checkout_date, adults, children,
                        total_estimated, source, status, notes, created_by)
                   values ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,'WALK-IN','CHECKED_IN',$11,$12)
                   returning id""",
                tid, code, guest_id, db_pg.a_uuid(habitacion["room_type_id"]), hid,
                today, checkout_date, adults, children, total_estimated, notes, quien,
            )
        except db_pg.ExclusionViolationError:
            raise HTTPException(
                status_code=409,
                detail="La habitación ya tiene una reserva en esas fechas"
            )

        stay_id = await conn.fetchval(
            """insert into stays (tenant_id, reservation_id, guest_id, room_id,
                                  checkin_at, status, created_by)
               values ($1, $2, $3, $4, now(), 'OPEN', $5)
               returning id""",
            tid, reservation_id, guest_id, hid, quien,
        )
        folio_id = await conn.fetchval(
            """insert into folios (tenant_id, reservation_id, stay_id, status)
               values ($1, $2, $3, 'OPEN')
               returning id""",
            tid, reservation_id, stay_id,
        )
        await conn.execute(
            "update rooms set occupancy_status = 'OCCUPIED' where id = $1", hid
        )

        await create_audit_log(conn, tid, user["user_id"], "reservation", "WALKIN",
                               None, {"code": code, "room": habitacion["number"]})

    reservation_id = str(reservation_id)
    stay_id = str(stay_id)
    folio_id = str(folio_id)

    return {
        "id": reservation_id,
        "code": code,
        "stay_id": stay_id,
        "folio_id": folio_id,
        "message": "Walk-in registrado exitosamente"
    }

# ============== EMAIL NOTIFICATIONS ==============
async def send_email_async(to_email: str, subject: str, html_content: str,
                           texto: str = ""):
    """Envia un correo por SMTP.

    Antes usaba la API de Resend, pero nunca llego a funcionar: RESEND_API_KEY
    no estaba definida y la funcion devolvia "skipped", asi que los correos se
    perdian en silencio. El correo de la casa es soporte@sisac.pe en cPanel, el
    mismo que usan FletePro y LicitaPro, de modo que los tres salen ahora por
    un solo buzon y una sola clave.
    """
    import smtplib
    from email.message import EmailMessage

    host = os.environ.get("SMTP_HOST")
    puerto = int(os.environ.get("SMTP_PORT", "465"))
    usuario = os.environ.get("SMTP_USER")
    cifrada = os.environ.get("SMTP_PASSWORD_B64")
    clave = base64.b64decode(cifrada).decode() if cifrada else os.environ.get("SMTP_PASSWORD", "")
    remitente = os.environ.get("SMTP_REMITENTE") or usuario

    if not (host and usuario and clave):
        # Ruidoso a proposito. La version anterior devolvia "skipped" cuando le
        # faltaba configuracion, o sea que respondia como si todo estuviera bien
        # y el correo no salia. Un fallo de envio tiene que verse en el log.
        logger.error(
            "SMTP sin configurar (faltan SMTP_HOST / SMTP_USER / SMTP_PASSWORD_B64): "
            "correo NO enviado a %s", to_email
        )
        return {"status": "error", "error": "SMTP no configurado"}

    def _enviar():
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = remitente
        msg["To"] = to_email
        # El texto plano va PRIMERO: en multipart/alternative el cliente elige
        # la ULTIMA parte que sabe pintar, asi que el HTML tiene que ir al
        # final. Sin la parte de texto, ademas, los filtros puntuan peor.
        msg.set_content(texto or "Este mensaje necesita un lector de correo con HTML.")
        msg.add_alternative(html_content, subtype="html")
        if puerto == 465:
            with smtplib.SMTP_SSL(host, puerto, timeout=30) as s:
                s.login(usuario, clave)
                s.send_message(msg)
        else:
            with smtplib.SMTP(host, puerto, timeout=30) as s:
                s.starttls()
                s.login(usuario, clave)
                s.send_message(msg)

    try:
        # En un hilo aparte: smtplib es bloqueante y esto corre dentro del loop.
        await asyncio.to_thread(_enviar)
        return {"status": "sent"}
    except Exception as e:
        logger.error("Fallo el envio de correo a %s: %s", to_email, e)
        return {"status": "error", "error": str(e)}

def generate_email_template(template_type: str, data: dict) -> tuple:
    """Arma (asunto, html, texto) de un correo a partir de su plantilla.

    Devuelve tambien la version en texto plano, no solo el HTML: hay clientes
    que no pintan HTML y los filtros de spam penalizan los mensajes que llegan
    en un solo formato.

    El diseno vive entero en plantillas_correo. Aqui solo se decide QUE dice
    cada correo; COMO se ve es cosa de alli, para que los cuatro cambien a la
    vez cuando cambie la marca.
    """
    hotel = data.get("tenant_name") or "Hotel"
    huesped = data.get("guest_name") or "Huésped"

    def soles(v):
        # El total puede llegar como Decimal, str o None segun quien llame.
        try:
            return "S/ %.2f" % float(v or 0)
        except (TypeError, ValueError):
            return "S/ —"

    if template_type == "RESERVATION_CONFIRMATION":
        asunto = "Confirmación de reserva %s — %s" % (data.get("code", ""), hotel)
        kw = dict(
            titulo="Tu reserva está confirmada",
            hotel=hotel,
            preencabezado="Código %s · Entrada %s" % (data.get("code", ""),
                                                      data.get("checkin_date", "")),
            intro=["Estimado/a %s," % huesped,
                   "Confirmamos tu reserva. Estos son los datos; guárdalos "
                   "para el día de la llegada."],
            filas=[
                ("Código de reserva", data.get("code") or "—"),
                ("Entrada", "%s · desde las 14:00" % (data.get("checkin_date") or "—")),
                ("Salida", "%s · hasta las 12:00" % (data.get("checkout_date") or "—")),
                ("Habitación", data.get("room_type") or "—"),
                ("Total estimado", soles(data.get("total"))),
            ],
            cierre=["Si necesitas cambiar algo, responde a este correo antes "
                    "de la fecha de entrada."],
            aviso="El total es estimado: no incluye consumos ni servicios "
                  "adicionales que se registren durante la estadía.",
        )

    elif template_type == "CHECKIN_CONFIRMATION":
        asunto = "Bienvenido a %s — check-in confirmado" % hotel
        kw = dict(
            titulo="Bienvenido, ya estás registrado",
            hotel=hotel,
            preencabezado="Habitación %s · Salida %s" % (
                data.get("room_number", ""), data.get("checkout_date", "")),
            intro=["Estimado/a %s," % huesped,
                   "Tu check-in quedó registrado. Que disfrutes la estadía."],
            filas=[
                ("Habitación", data.get("room_number") or "—"),
                ("Piso", data.get("floor") or "—"),
                ("Salida", "%s · hasta las 12:00" % (data.get("checkout_date") or "—")),
            ],
            cierre=["Cualquier cosa que necesites, recepción está a tu "
                    "disposición."],
        )

    elif template_type == "CHECKOUT_REMINDER":
        asunto = "Recordatorio de salida — %s" % hotel
        kw = dict(
            titulo="Se acerca tu salida",
            hotel=hotel,
            preencabezado="Salida %s hasta las 12:00" % data.get("checkout_date", ""),
            intro=["Estimado/a %s," % huesped,
                   "Te recordamos la hora de salida para que organices el día."],
            filas=[
                ("Salida", "%s · hasta las 12:00" % (data.get("checkout_date") or "—")),
                ("Habitación", data.get("room_number") or "—"),
            ],
            cierre=["Si necesitas salida tardía, consúltalo en recepción: "
                    "depende de la disponibilidad del día."],
        )

    else:  # PAYMENT_RECEIPT
        asunto = "Comprobante %s — %s" % (data.get("invoice_number", ""), hotel)
        kw = dict(
            titulo="Comprobante de pago",
            hotel=hotel,
            preencabezado="%s %s" % (data.get("invoice_type") or "Comprobante",
                                     data.get("invoice_number", "")),
            intro=["Estimado/a %s," % (data.get("client_name") or huesped),
                   "Registramos tu pago. Este es el detalle."],
            filas=[
                ("Comprobante", data.get("invoice_number") or "—"),
                ("Tipo", data.get("invoice_type") or "—"),
                ("Fecha", data.get("date") or "—"),
                ("Total pagado", soles(data.get("total"))),
            ],
            cierre=["Gracias por tu preferencia."],
        )

    texto, html = plantillas_correo.componer(**kw)
    return asunto, html, texto

class NotificationRequest(BaseModel):
    template: EmailTemplate
    recipient_email: EmailStr
    data: dict

@api_router.post("/notifications/send")
async def send_notification(
    request: NotificationRequest,
    user: dict = Depends(get_current_user)
):
    """Envia un correo a partir de una plantilla."""
    tid = db_pg.a_uuid(user["tenant_id"], "tenant_id")

    async with db_pg.tx(user) as conn:
        tenant = await db_pg.uno(
            conn, "select name, nombre_comercial from tenants where id = $1", tid
        )

    data = request.data.copy()
    data["tenant_name"] = (tenant.get("nombre_comercial") or tenant.get("name") or "Hotel") if tenant else "Hotel"

    subject, html, texto = generate_email_template(request.template.value, data)
    # El envio va FUERA de la transaccion: mandar un correo no se puede
    # deshacer, asi que no tiene sentido tener la base bloqueada esperandolo.
    result = await send_email_async(request.recipient_email, subject, html, texto)

    async with db_pg.tx(user) as conn:
        await conn.execute(
            """insert into notification_logs (tenant_id, template, recipient, status)
               values ($1, $2, $3, $4)""",
            tid, request.template.value, request.recipient_email, result.get("status"),
        )

    return result

@api_router.get("/notifications/logs")
async def get_notification_logs(
    limit: int = Query(50, ge=1, le=200),
    user: dict = Depends(require_roles(Role.ADMIN))
):
    """Historial de correos enviados."""
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        return await db_pg.varias(
            conn,
            """select * from notification_logs
               where ($1::uuid is null or tenant_id = $1)
               order by created_at desc
               limit $2""",
            tid, limit,
        )

# ============== CALENDAR DATA ENDPOINT ==============
@api_router.get("/calendar/reservations")
async def get_calendar_reservations(
    start_date: date = Query(...),
    end_date: date = Query(...),
    user: dict = Depends(get_current_user)
):
    """Reservas del calendario, con soporte de arrastrar y soltar."""
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        # Una consulta en lugar de 1 + 3N. Antes, por CADA reserva del rango se
        # pedian por separado la habitacion, el huesped y el tipo: un mes con
        # 100 reservas eran 301 viajes a la base cada vez que alguien abria el
        # calendario.
        #
        # Ademas el solapamiento se calcula bien: la condicion es "empieza antes
        # de que acabe el rango Y termina despues de que empiece". La version
        # anterior enumeraba tres casos con OR y se dejaba fuera reservas que
        # cruzaban el rango por un solo dia.
        filas = await db_pg.varias(
            conn,
            """select r.id, r.code, r.room_id, r.status,
                      coalesce(g.full_name, 'Sin huésped') as title,
                      h.number as room_number,
                      rt.name  as room_type,
                      r.checkin_date  as start,
                      r.checkout_date as "end",
                      case when r.status = 'CONFIRMED' then '#3B82F6' else '#10B981' end as color
               from reservations r
               left join guests     g  on g.id  = r.guest_id
               left join rooms      h  on h.id  = r.room_id
               left join room_types rt on rt.id = r.room_type_id
               where ($1::uuid is null or r.tenant_id = $1)
                 and r.status in ('CONFIRMED', 'CHECKED_IN')
                 and r.checkin_date  <= $3
                 and r.checkout_date >= $2
               order by r.checkin_date""",
            tid, start_date, end_date,
        )
    result = filas
    
    return result

@api_router.put("/calendar/reservations/{reservation_id}/move")
async def move_reservation(
    reservation_id: str,
    new_room_id: str = Body(...),
    new_checkin: date = Body(None),
    new_checkout: date = Body(None),
    user: dict = Depends(get_current_user)
):
    """Mueve una reserva de habitacion o de fechas (arrastrar y soltar)."""
    tid = tenant_de(user)
    rid = id_valido(reservation_id, "reservation_id")
    nueva = id_valido(new_room_id, "new_room_id")

    async with db_pg.tx(user) as conn:
        anterior = await db_pg.uno(
            conn,
            """select tenant_id, room_id, checkin_date, checkout_date
               from reservations
               where id = $1 and ($2::uuid is null or tenant_id = $2)
               for update""",
            rid, tid,
        )
        if not anterior:
            raise HTTPException(status_code=404, detail="Reserva no encontrada")

        # La habitacion destino tiene que ser del MISMO hotel. La consulta de
        # conflicto anterior no filtraba por tenant_id: buscaba cualquier
        # reserva del sistema con ese room_id.
        if not await conn.fetchval(
            "select 1 from rooms where id = $1 and tenant_id = $2",
            nueva, db_pg.a_uuid(anterior["tenant_id"]),
        ):
            raise HTTPException(status_code=404, detail="Habitación no encontrada")

        try:
            # Ya no hay consulta previa de solapamiento: la decide el constraint
            # de db/migrations/001, que no tiene ventana de carrera.
            movida = await db_pg.uno(
                conn,
                """update reservations set
                       room_id       = $3,
                       checkin_date  = coalesce($4::date, checkin_date),
                       checkout_date = coalesce($5::date, checkout_date)
                   where id = $1 and ($2::uuid is null or tenant_id = $2)
                   returning checkin_date, checkout_date""",
                rid, tid, nueva, new_checkin, new_checkout,
            )
        except db_pg.ExclusionViolationError:
            raise HTTPException(status_code=400, detail="Conflicto con otra reserva en esa habitación")
        except db_pg.CheckViolationError:
            raise HTTPException(status_code=400, detail="La fecha de salida debe ser posterior a la de entrada")

        await create_audit_log(
            conn, anterior["tenant_id"], user["user_id"], "reservation", "MOVE",
            {"room_id": anterior.get("room_id")},
            {"room_id": new_room_id,
             "dates": f"{movida['checkin_date']} - {movida['checkout_date']}"},
        )

    return {"message": "Reserva movida exitosamente"}

# ============== SEARCH ENDPOINT ==============
@api_router.get("/search")
async def global_search(q: str = Query(..., min_length=2), user: dict = Depends(get_current_user)):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        # Las tres busquedas comparten una sola ida a la base. El texto viaja
        # como parametro: antes se inyectaba crudo en un $regex, asi que una
        # busqueda como "(a+)+" podia dejar la consulta girando sobre toda la
        # tabla desde la caja de busqueda de la interfaz.
        fila = await db_pg.uno(
            conn,
            """select
                   coalesce((select jsonb_agg(to_jsonb(g.*))
                             from (select * from guests
                                   where ($1::uuid is null or tenant_id = $1)
                                     and (full_name  ilike '%' || $2 || '%'
                                       or doc_number ilike '%' || $2 || '%'
                                       or email      ilike '%' || $2 || '%')
                                   order by full_name limit 10) g), '[]'::jsonb) as guests,
                   coalesce((select jsonb_agg(to_jsonb(r.*))
                             from (select * from reservations
                                   where ($1::uuid is null or tenant_id = $1)
                                     and code ilike '%' || $2 || '%'
                                   order by checkin_date desc limit 10) r), '[]'::jsonb) as reservations,
                   coalesce((select jsonb_agg(to_jsonb(h.*))
                             from (select * from rooms
                                   where ($1::uuid is null or tenant_id = $1)
                                     and number ilike '%' || $2 || '%'
                                   order by number limit 10) h), '[]'::jsonb) as rooms""",
            tid, q,
        )
    return fila

# ============== AUDIT LOG ENDPOINT ==============
@api_router.get("/audit-logs")
async def list_audit_logs(
    entity: Optional[str] = None,
    action: Optional[str] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    user: dict = Depends(require_roles(Role.ADMIN, Role.SUPER_ADMIN))
):
    tid = tenant_de(user)
    async with db_pg.tx(user) as conn:
        return await db_pg.varias(
            conn,
            """select * from audit_logs
               where ($1::uuid is null or tenant_id = $1)
                 and ($2::text is null or entity = $2)
                 and ($3::text is null or action = $3)
                 and ($4::date is null or created_at >= $4::date)
                 and ($5::date is null or created_at < ($5::date + 1))
               order by created_at desc
               limit 500""",
            tid, entity, action, from_date, to_date,
        )

# ============== SEED DATA ENDPOINT ==============
@api_router.post("/seed")
async def seed_demo_data(user: dict = Depends(require_roles(Role.SUPER_ADMIN))):
    """Crea un hotel de demostracion con datos de ejemplo."""
    async with db_pg.tx_global("crear el hotel demo: se crea el tenant desde cero") as conn:
        existente = await conn.fetchval("select id from tenants where ruc = $1", "20123456789")
        if existente:
            return {"message": "Datos demo ya existen", "tenant_id": str(existente)}

        tenant_id = await conn.fetchval(
            """insert into tenants (name, ruc, razon_social, nombre_comercial,
                                    address, ubigeo, phone, email)
               values ('Hotel Demo', '20123456789', 'Hotel Demo S.A.C.', 'Hotel Demo',
                       'Av. Principal 123, Lima, Perú', '150101',
                       '+51 1 234 5678', 'demo@hoteldemo.com')
               returning id"""
        )

        # Las contrasenas demo salen de una variable de entorno. Antes estaban
        # escritas en el codigo ("admin123"), asi que cualquiera que leyera el
        # repositorio -- publico o no -- tenia las credenciales de todo hotel
        # que hubiera corrido este endpoint.
        clave_demo = os.environ.get("SEED_DEMO_PASSWORD") or secrets.token_urlsafe(12)
        cuentas = [
            ("admin@demo.com",     "Admin Demo",     "ADMIN"),
            ("recepcion@demo.com", "María García",   "RECEPTIONIST"),
            ("limpieza@demo.com",  "Carlos López",   "HOUSEKEEPING"),
            ("seguridad@demo.com", "Pedro Ramirez",  "SECURITY"),
        ]
        hash_demo = hash_password(clave_demo)
        await conn.executemany(
            """insert into users (tenant_id, email, password_hash, full_name, role)
               values ($1, $2, $3, $4, $5::user_role)""",
            [(tenant_id, email, hash_demo, nombre, rol) for email, nombre, rol in cuentas],
        )

        tipos = await conn.fetch(
            """insert into room_types (tenant_id, name, capacity, amenities, base_price)
               values ($1, 'Estándar', 2, $2, 150.00),
                      ($1, 'Superior', 2, $3, 220.00),
                      ($1, 'Suite',    4, $4, 350.00)
               returning id""",
            tenant_id,
            ["WiFi", "TV", "Aire Acondicionado"],
            ["WiFi", "TV", "Aire Acondicionado", "Minibar", "Caja Fuerte"],
            ["WiFi", "TV", "Aire Acondicionado", "Minibar", "Caja Fuerte", "Jacuzzi", "Sala"],
        )
        rt_ids = [r["id"] for r in tipos]

        habitaciones = []
        for piso in range(1, 4):
            for num in range(1, 11):
                idx = 0 if num <= 6 else (1 if num <= 9 else 2)
                habitaciones.append((tenant_id, rt_ids[idx], f"{piso}{num:02d}", piso))
        await conn.executemany(
            """insert into rooms (tenant_id, room_type_id, number, floor)
               values ($1, $2, $3, $4)""",
            habitaciones,
        )

        productos = [
            ("Minibar - Agua",        "MINIBAR",    5.00),
            ("Minibar - Gaseosa",     "MINIBAR",    8.00),
            ("Minibar - Cerveza",     "MINIBAR",   12.00),
            ("Lavandería - Camisa",   "LAVANDERIA", 15.00),
            ("Lavandería - Pantalón", "LAVANDERIA", 18.00),
            ("Late Checkout",         "SERVICIOS",  50.00),
            ("Early Checkin",         "SERVICIOS",  50.00),
            ("Daño - Toalla",         "DANOS",      25.00),
        ]
        await conn.executemany(
            """insert into products (tenant_id, name, category, unit_price)
               values ($1, $2, $3, $4)""",
            [(tenant_id, n, c, p) for n, c, p in productos],
        )

    return {
        "message": "Datos demo creados exitosamente",
        "tenant_id": str(tenant_id),
        "usuarios": [c[0] for c in cuentas],
        "password": clave_demo,
        "aviso": "Esta contraseña se muestra una sola vez. Cámbiala tras el primer acceso."
    }

# Solo funciona con la base vacia: crea el primer SUPER_ADMIN.
@api_router.post("/setup")
async def initial_setup():
    """Crea el super administrador inicial. Solo si no existe ningun usuario."""
    # La contrasena ya NO esta escrita en el codigo. Antes era "superadmin123",
    # y como la base nueva arranca vacia eso era una carrera real: el primero
    # que llamara a este endpoint despues del despliegue -- cualquiera en
    # internet -- se quedaba con el superadministrador del sistema.
    #
    # Ahora sale de SETUP_PASSWORD si esta definida y, si no, se genera una al
    # azar que se devuelve UNA sola vez en esta respuesta.
    email = os.environ.get("SETUP_EMAIL", "superadmin@sistema.com")
    password = os.environ.get("SETUP_PASSWORD") or secrets.token_urlsafe(16)

    # La condicion "solo si no hay usuarios" vive dentro de la funcion de
    # Postgres, con un lock de tabla: dos llamadas simultaneas no pueden crear
    # dos superadmins.
    nuevo_id = await db_pg.setup_inicial(email, hash_password(password), "Super Administrador")
    if nuevo_id is None:
        raise HTTPException(
            status_code=400,
            detail="Sistema ya inicializado. Use /seed con credenciales de Super Admin"
        )

    return {
        "message": "Super Admin creado exitosamente",
        "credentials": {"email": email, "password": password},
        "aviso": "Esta contraseña se muestra una sola vez. Cámbiala tras el primer acceso."
    }

# Root endpoint
@api_router.get("/")
async def root():
    return {"message": "ZenStay API v1.0", "status": "running"}

# Health check
@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# Include router
app.include_router(api_router)

# La compra publica: /precios, /comprar/{plan} y /comprar/pedido/{numero}.
#
# Va DESPUES de api_router y ANTES del catch-all de abajo, y el orden es lo
# unico que hace que funcione: el catch-all `@app.get("/{ruta:path}")` atrapa
# cualquier cosa, asi que un router montado despues de el no recibiria nunca
# una peticion. Estas paginas se construyen enteras en el servidor porque el
# validador de Izipay descarga el HTML y puede no ejecutar JavaScript (ver la
# cabecera de backend/checkout.py).
app.include_router(checkout.router)

# CORS. Los origenes se validan al arrancar (ver arriba): la lista nunca es '*',
# porque las peticiones van con credenciales y esa combinacion la rechazan los
# navegadores ademas de ser insegura.
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=CORS_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------------
# La SPA
# ---------------------------------------------------------------------------
# El backend sirve tambien el frontend compilado, igual que FletePro y
# LicitaPro en este mismo VPS: asi el dominio del tunel apunta a un solo puerto
# y no hace falta ningun proxy delante.
#
# Esto es lo que faltaba: sin este bloque, GET / devolvia el 404 de FastAPI y la
# aplicacion no se veia por ningun lado, aunque /api respondiera bien.
#
# Va DESPUES de include_router a proposito. El orden importa: las rutas /api ya
# estan registradas, asi que el catch-all de abajo solo atrapa lo que no sea de
# la API.
FRONTEND_BUILD = ROOT_DIR.parent / "frontend" / "build"
LANDING = ROOT_DIR / "landing"

# La landing es HTML estatico y no parte de la SPA a proposito. Es la pagina
# que tiene que vender: carga al instante, la lee un buscador, y no obliga a
# descargar el megabyte de JavaScript del sistema a alguien que todavia no sabe
# si le interesa. El sistema vive en la raiz: /login, /dashboard, /reservations...
if LANDING.exists():
    app.mount("/landing-assets", StaticFiles(directory=LANDING), name="landing_assets")


def _version_css() -> str:
    """Huella del CSS de la landing, para invalidar la cache de Cloudflare.

    Sin esto, un despliegue no se ve: Cloudflare cachea los estaticos con
    max-age de 4 horas, asi que el navegador seguia recibiendo el CSS anterior
    -- `cf-cache-status: HIT` -- mientras el contenedor ya servia el nuevo. La
    pagina quedaba con los colores viejos y nada en los logs indicaba el
    porque.

    Se calcula al vuelo y no una vez al arrancar: el contenedor se reconstruye
    en cada despliegue, asi que da igual en cuanto a coste, y en desarrollo
    permite editar el CSS sin reiniciar.

    Delega en checkout.version_css() y no calcula lo suyo: las paginas de la
    compra cargan ESE MISMO archivo, y si cada modulo sacara su propia huella
    un despliegue podria dejar el checkout apuntando a una version y la landing
    a otra.
    """
    return checkout.version_css()


def _pagina_landing(archivo: Path) -> HTMLResponse:
    """Sirve una pagina de la landing con el CSS versionado.

    El HTML va con `no-cache` a proposito: es lo que apunta a la version nueva
    del CSS, asi que cachearlo dejaria clavada la referencia vieja y no habria
    servido de nada versionar nada.
    """
    html = archivo.read_text(encoding="utf-8").replace(
        "/landing-assets/estilo.css", f"/landing-assets/estilo.css?v={_version_css()}"
    )
    return HTMLResponse(html, headers={"Cache-Control": "no-cache, must-revalidate"})

if FRONTEND_BUILD.exists():
    # Los assets con hash en el nombre (JS, CSS). package.json declara
    # homepage="/", asi que el index.html pide /static/js/main.<hash>.js.
    app.mount("/static", StaticFiles(directory=FRONTEND_BUILD / "static"), name="static")


# Cuanto puede guardarse cada cosa del build.
#
# Tres politicas, porque el riesgo de cada fichero es distinto. La primera es
# la que faltaba y la que rompia los despliegues.
_UN_ANO_SPA = 60 * 60 * 24 * 365
_UN_DIA_SPA = 60 * 60 * 24


def _cache_del_build(ruta: Path) -> str:
    """Devuelve el Cache-Control que le toca a un fichero del build de la SPA."""
    nombre = ruta.name

    # index.html NUNCA en duro: es el que NOMBRA al bundle. Si el navegador se
    # queda con el viejo sigue mostrando la version anterior de la aplicacion,
    # indefinidamente. Sin cabecera pasaba justo eso: el navegador aplicaba su
    # heuristica -un 10% del tiempo desde Last-Modified- y se quedaba con el.
    #
    # El service worker va igual y por lo mismo: mientras el navegador lo tenga
    # cacheado no se entera de que hay uno nuevo, y el service worker es quien
    # manda sobre lo que acaba viendo el usuario.
    #
    # s-maxage=0 ademas de no-cache: Cloudflare tiene el "Browser Cache TTL"
    # del panel en 4 horas y REESCRIBE la cabecera que sale de aqui para los
    # ficheros que cachea en el borde. Medido: el origen mandaba
    # 'no-cache, must-revalidate' y al navegador le llegaba 'max-age=14400'.
    # Con s-maxage=0 el borde no lo retiene y la cabecera del origen sobrevive.
    if nombre in ("index.html", "service-worker.js"):
        return "public, no-cache, must-revalidate, s-maxage=0"

    # Lo que lleva hash de contenido en el nombre es inmutable por definicion.
    if re.search(r"\.[0-9a-f]{8,}\.[a-z0-9]+$", nombre):
        return f"public, max-age={_UN_ANO_SPA}, immutable"

    # El resto -logo, favicon, manifest- conserva el nombre entre despliegues.
    return f"public, max-age={_UN_DIA_SPA}, stale-while-revalidate={_UN_DIA_SPA * 7}"


@app.get("/{ruta:path}")
async def servir_web(ruta: str, request: Request):
    """Reparte entre la landing y la aplicacion.

        /            -> landing (comercial)
        /privacidad, /terminos, /reclamaciones -> landing
        /app/...     -> 301 a la misma ruta sin /app (enlaces y marcadores viejos)
        /algo.ext    -> archivo del build si existe (favicon, logos, manifest)
        lo demas     -> la SPA (/login, /registro, /dashboard, /reservations...)

    El catch-all va DESPUES de include_router, asi que /api ya esta resuelto y
    no pasa por aqui.
    """
    # 0. La aplicacion vivio bajo /app hasta septiembre de 2026. Los
    #    recepcionistas tienen /app/dashboard en marcadores y hay enlaces por
    #    correo: un 301 los lleva a la ruta nueva conservando la consulta.
    if ruta == "app" or ruta.startswith("app/"):
        destino = "/" + ruta[4:]
        if request.url.query:
            destino += "?" + request.url.query
        return RedirectResponse(destino, status_code=301)

    # 1. Paginas de la landing.
    #
    #    /precios YA NO esta aqui ni redirige al ancla /#planes: ahora es una
    #    pagina de verdad, construida en el servidor por backend/checkout.py y
    #    montada antes que este catch-all. El ancla no servia como catalogo
    #    para nadie de fuera -- ni para la pasarela, ni para un buscador, ni
    #    para un enlace por WhatsApp -- porque los precios de la portada los
    #    pinta un fetch desde el navegador.
    if LANDING.exists():
        # /registro ya NO esta aqui: el alta es una pantalla de la SPA, con la
        # misma concha que el login, y la resuelve React.
        paginas = {
            "": "index.html",
            "terminos": "terminos.html",
            "privacidad": "privacidad.html",
            "reclamaciones": "reclamaciones.html",
        }
        if ruta in paginas:
            archivo = LANDING / paginas[ruta]
            if archivo.is_file():
                return _pagina_landing(archivo)

    if not FRONTEND_BUILD.exists():
        raise HTTPException(status_code=404, detail="No encontrado")

    # 2. Un archivo real del build (logo.png, manifest.json, favicon...).
    #
    #    resolve() y el chequeo de prefijo evitan que una ruta con ../ se escape
    #    del directorio del build y sirva archivos del contenedor.
    if ruta:
        candidato = FRONTEND_BUILD / ruta
        if candidato.is_file() and str(candidato.resolve()).startswith(str(FRONTEND_BUILD.resolve())):
            return FileResponse(
                candidato, headers={"Cache-Control": _cache_del_build(candidato)}
            )

    # 3. Si la ruta PARECE un archivo (tiene extension) y no se encontro, es un
    #    404 de verdad -- no una ruta de la SPA.
    #
    #    Importa mas de lo que parece: devolver el index.html con un 200 para
    #    /app/logo-zenstay.png hizo que Cloudflare cachease ese HTML como si
    #    fuera la imagen, con max-age de 4 horas. El logotipo salio roto en todo
    #    el sistema y siguio roto despues de arreglar la ruta, porque lo que
    #    servia el CDN ya no dependia del servidor. Un 404 no se cachea asi.
    if "." in ruta.rsplit("/", 1)[-1]:
        raise HTTPException(status_code=404, detail="No encontrado")

    # 4. Cualquier otra cosa es una ruta de la SPA. Se devuelve el index.html y
    #    React resuelve el enrutado: es lo que hace falta para que recargar
    #    /reservations o entrar por un enlace directo no de 404.
    indice = FRONTEND_BUILD / "index.html"
    return FileResponse(indice, headers={"Cache-Control": _cache_del_build(indice)})


if not FRONTEND_BUILD.exists():
    logger.warning(
        "No existe %s: el backend sirve solo la API. En el VPS el build llega "
        "dentro de la imagen (ver backend/Dockerfile).",
        FRONTEND_BUILD,
    )


@app.on_event("shutdown")
async def cerrar_conexiones():
    await db_pg.close_pool()
